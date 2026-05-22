"""
Конвертер: regel_verben-presens.xlsx → блок REGEL_VERBS в data.js

Запуск: python update_regel_verbs.py

Что делает:
1. Читает все строки из xlsx (verb, ru, ich, du, er/sie/es, wir, ihr, sie/Sie)
2. Перенумеровывает id с нуля (r001, r002, ...)
3. Вставляет/перезаписывает блок `const REGEL_VERBS = [...]` в data.js
4. Сохраняет data.js обратно

ID связки в data.js: по полю `verb` (текст инфинитива).
ID из xlsx игнорируются — они только для удобства редактирования xlsx.
"""
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Нужен openpyxl. Установи: pip install openpyxl")
    sys.exit(1)

# ─── Пути ─────────────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).parent
XLSX_PATH = SCRIPT_DIR / "regel_verben-presens.xlsx"
DATA_JS = SCRIPT_DIR / "data.js"

# ─── Чтение xlsx ──────────────────────────────────────────────────────
print(f"Читаю {XLSX_PATH}...")
wb = load_workbook(XLSX_PATH)
ws = wb.active  # первый лист

regel_verbs = []
seen = set()  # для защиты от дублей
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row or not row[1]:
        continue
    verb = str(row[1]).strip()
    if verb in seen:
        print(f"  ⚠ Дубль: {verb} — пропускаю")
        continue
    seen.add(verb)
    ru = str(row[2]).strip() if row[2] else ""
    forms = [str(row[i]).strip() if row[i] else "" for i in (3, 4, 5, 6, 7, 8)]
    if any(not f for f in forms):
        print(f"  ⚠ {verb}: неполные формы, пропускаю")
        continue
    regel_verbs.append({
        "verb": verb,
        "ru": ru,
        "forms": forms,
    })

print(f"✓ Прочитано: {len(regel_verbs)} глаголов")

# ─── Сериализация в JS ────────────────────────────────────────────────
def js_str(s):
    s = str(s).replace("\\", "\\\\").replace('"', '\\"')
    return f'"{s}"'

PRONOUNS = ["ich", "du", "er/sie/es", "wir", "ihr", "sie/Sie"]

lines = []
lines.append("// ═══════════════════════════════════════════════════════════════")
lines.append(f"// REGEL_VERBS — таблица regelmäßig спряжений в Präsens ({len(regel_verbs)} глаголов)")
lines.append("// Источник: regel_verben-presens.xlsx (импорт скриптом update_regel_verbs.py)")
lines.append(f"// Местоимения: {', '.join(PRONOUNS)}")
lines.append("// Связь с VOCAB — по полю `verb` (текст инфинитива)")
lines.append("// ═══════════════════════════════════════════════════════════════")
lines.append("const REGEL_VERBS = [")
for i, v in enumerate(regel_verbs, 1):
    vid = f"r{i:03d}"
    forms_js = ", ".join(js_str(f) for f in v["forms"])
    lines.append(f'  {{ id: {js_str(vid)}, verb: {js_str(v["verb"])}, '
                 f'ru: {js_str(v["ru"])}, forms: [{forms_js}] }},')
lines.append("];")
lines.append("")

new_block = "\n".join(lines)

# ─── Вставка в data.js ────────────────────────────────────────────────
print(f"Читаю {DATA_JS}...")
content = DATA_JS.read_text(encoding="utf-8")

# Маркеры начала и конца блока (комментарий + const + закрывающая ])
# Стратегия: если в data.js уже есть `const REGEL_VERBS = [...]` — заменяем
# Если нет — вставляем перед `if (typeof module ...)` (или в конец)
pattern_block = re.compile(
    r"// ═{10,}\s*\n// REGEL_VERBS[\s\S]*?^const REGEL_VERBS = \[[\s\S]*?^\];\s*\n",
    re.MULTILINE
)
if pattern_block.search(content):
    print("  Нашёл существующий блок — заменяю")
    content = pattern_block.sub(new_block + "\n", content, count=1)
else:
    print("  Нового блока ещё нет — вставляю перед module.exports")
    inject_marker = "if (typeof module !== 'undefined' && module.exports) {"
    if inject_marker in content:
        content = content.replace(inject_marker, new_block + "\n" + inject_marker, 1)
    else:
        # Просто в конец файла
        content = content.rstrip() + "\n\n" + new_block + "\n"

# Обновить экспорт чтобы REGEL_VERBS экспортировалось
# Ищем module.exports = { ... } и добавляем REGEL_VERBS если его там нет
export_pattern = re.compile(r"module\.exports = \{ ([^}]*) \};")
m = export_pattern.search(content)
if m:
    exports = m.group(1)
    if "REGEL_VERBS" not in exports:
        new_exports = exports.rstrip(", ") + ", REGEL_VERBS"
        content = export_pattern.sub(f"module.exports = {{ {new_exports} }};", content, count=1)
        print("  Добавил REGEL_VERBS в module.exports")

DATA_JS.write_text(content, encoding="utf-8")
print(f"✓ Обновлён {DATA_JS}")
print(f"  Размер: {DATA_JS.stat().st_size:,} байт")
