"""
update_data.py — собирает data.js из database.xlsx

Запуск: python update_data.py

Что делает:
1. Читает все листы database.xlsx (nouns, verbs, adjectives, adverbs, pronouns,
   numbers, phrases, sounds, terms, rules, questions) + лист Settings-XLSX.
2. Для verbs: regel без форм → автогенератор; regel с формами → ручные.
   Для unregel → берёт формы из ячеек.
3. Сохраняет BLOCKS, TOPIC_TITLES, TAB_TITLES, PHRASE_UNITS, SENTENCE_TEMPLATES
   из существующего data.js (не пересоздаёт их — это код фронта, не контент).
4. Settings-XLSX → const TAXONOMY (двухуровневая иерархия domen → group).
5. Финальный файл data.js пишется заново.

СХЕМА (новая, domen/group):
- topic заменён на пару domen + group на всех контентных листах.
- Settings-XLSX — мастер-список валидных пар (page, domen, group, …).
- Глаголы: de / er_sie_es / sie_Sie + флаги separable, prefix, reflexive,
  impersonal, case, praeteritum.
- Поля example_de / example_ru + quiz_use (годится ли ПРИМЕР для теста-сборки).

ПРИНЦИПЫ:
- Читаются ВСЕ колонки схемы (даже сейчас пустые). Маппинг объявлен в SCHEMA —
  заполнишь колонку в xlsx, и она поедет в data.js без правок кода.
- ID берутся из xlsx как есть (уважаются). Пустой id → автоген (prefix + номер) + варн.
  Дубликаты id внутри листа → варн.
- Связь между сущностями — по тексту слова (`de`).
- Тип вопроса build трактуется как алиас tiles (движок тренажёра знает только tiles).
- Колонка studied игнорируется (пережиток, удаляется из xlsx).
"""
import argparse
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Нужен openpyxl. pip install openpyxl")
    sys.exit(1)

# Прогресс DeepL-переводов (тысячи запросов, минуты между печатями) должен
# быть виден по мере выполнения, а не одним куском при выходе — иначе при
# перенаправлении в файл/лог stdout копится в блочном буфере до конца процесса.
try:
    sys.stdout.reconfigure(line_buffering=True)
except AttributeError:
    pass  # старый Python без reconfigure — переживём без live-вывода

SCRIPT_DIR = Path(__file__).parent
XLSX_PATH = SCRIPT_DIR / "database.xlsx"
OLD_DATA_JS = SCRIPT_DIR / "data.js"
OUT_DATA_JS = SCRIPT_DIR / "data.js"

# Колонки, которые сознательно игнорируются (пережитки)
IGNORED_COLS = {"studied"}

# ═══════════════════════════════════════════════════════════════
# АВТОГЕНЕРАТОР Präsens (regelmäßig)
# ═══════════════════════════════════════════════════════════════
def conj_regelmaessig(infinitive, pronoun):
    is_eln = infinitive.endswith("eln")
    is_ern = infinitive.endswith("ern")
    is_eln_or_rn = is_eln or is_ern

    if is_eln_or_rn:
        stem = infinitive[:-1]
    elif infinitive.endswith("en"):
        stem = infinitive[:-2]
    elif infinitive.endswith("n"):
        stem = infinitive[:-1]
    else:
        stem = infinitive

    ends_in_td = bool(re.search(r"[td]$", stem))
    needs_e_cluster = bool(re.search(r"(chn|ffn|tm|dm|gn)$", stem))
    needs_e = ends_in_td or needs_e_cluster
    sibilant_end = bool(re.search(r"[sßzx]$", stem))

    if pronoun == "ich":
        if is_eln: return stem[:-2] + "le"
        return stem + "e"
    if pronoun == "du":
        if sibilant_end: return stem + "t"
        if needs_e: return stem + "est"
        return stem + "st"
    if pronoun in ("er", "sie", "es"):
        if needs_e: return stem + "et"
        return stem + "t"
    if pronoun in ("wir", "Sie", "sie_pl"):
        if is_eln_or_rn: return stem + "n"
        return stem + "en"
    if pronoun == "ihr":
        if needs_e: return stem + "et"
        return stem + "t"
    return stem + "en"

UNREGELMAESSIG_VERBS = {
    "sein", "haben", "werden", "wissen", "mögen", "müssen", "können",
    "dürfen", "sollen", "wollen", "lassen", "fahren", "laufen", "schlafen",
    "lesen", "sehen", "geben", "nehmen", "kommen", "gehen", "stehen",
    "tragen", "waschen", "fallen", "halten", "schreiben",
}

def conj_all_forms(infinitive):
    """Возвращает 6 форм Präsens: [ich, du, er, wir, ihr, sie]"""
    if infinitive in UNREGELMAESSIG_VERBS:
        raise ValueError(
            f"conj_all_forms: «{infinitive}» — нерегулярный глагол, "
            f"автоген запрещён. Укажи тип unregel и формы вручную в xlsx."
        )
    return [conj_regelmaessig(infinitive, p)
            for p in ("ich", "du", "er", "wir", "ihr", "sie_pl")]

# ═══════════════════════════════════════════════════════════════
# АВТОГЕНЕРАТОР Partizip II (regelmäßig)
# ═══════════════════════════════════════════════════════════════
INSEPARABLE_PREFIXES = ("be", "ent", "er", "ge", "ver", "zer", "miss", "emp")
SEPARABLE_PREFIXES = (
    "ab", "an", "auf", "aus", "bei", "ein", "fest", "fort", "her", "hin",
    "los", "mit", "nach", "vor", "weg", "weiter", "zu", "zurück", "zusammen"
)

def partizip2_regelmaessig(infinitive):
    if infinitive.endswith("ieren"):
        return infinitive[:-2] + "t"

    if infinitive.endswith("en"):
        stem = infinitive[:-2]
    elif infinitive.endswith("n"):
        stem = infinitive[:-1]
    else:
        stem = infinitive

    needs_et = bool(re.search(r"[td]$", stem)) or \
               bool(re.search(r"(chn|ffn|tm|dm|gn)$", stem))
    ending = "et" if needs_et else "t"

    for prefix in INSEPARABLE_PREFIXES:
        if infinitive.startswith(prefix) and len(infinitive) > len(prefix) + 2:
            return stem + ending

    for prefix in SEPARABLE_PREFIXES:
        if infinitive.startswith(prefix) and len(infinitive) > len(prefix) + 2:
            sub_inf = infinitive[len(prefix):]
            sub_p2 = partizip2_regelmaessig(sub_inf)
            return prefix + sub_p2

    return "ge" + stem + ending

# ═══════════════════════════════════════════════════════════════
# СЕРИАЛИЗАЦИЯ В JS
# ═══════════════════════════════════════════════════════════════
def js_str(s):
    if s is None: return "null"
    s = (str(s).replace("\\", "\\\\").replace('"', '\\"')
         .replace("\n", "\\n").replace("\r", "\\r").replace("\t", "\\t"))
    return f'"{s}"'

def js_value(v):
    if v is None: return "null"
    if isinstance(v, bool): return "true" if v else "false"
    if isinstance(v, (int, float)): return str(v)
    if isinstance(v, list):
        return "[" + ", ".join(js_value(x) for x in v) + "]"
    if isinstance(v, dict):
        # Всегда квотим ключи — они могут содержать `:` и др. спецсимволы.
        return "{ " + ", ".join(f"{js_str(str(k))}: {js_value(val)}" for k, val in v.items()) + " }"
    return js_str(v)

def item_to_js(item, key_order=None):
    if key_order:
        keys = [k for k in key_order if k in item]
        # любые ключи не из key_order — в конец (чтобы новые поля не терялись)
        keys += [k for k in item if k not in key_order]
    else:
        keys = list(item.keys())
    parts = [f"{k}: {js_value(item[k])}" for k in keys]
    return "  { " + ", ".join(parts) + " },"

# ═══════════════════════════════════════════════════════════════
# СЛОЙ 2 МУЛЬТИЯЗЫЧНОСТИ (docs/en-fix/BRIEF-en-fix.md) — DE→EN через DeepL.
# Без DEEPL_API_KEY просто тихо пропускается — сайт от поля en не зависит,
# это доп. данные для переключателя языка контента (сейчас выключен,
# см. этап A: window.I18N_CONTENT_EN в i18n/i18n.js).
# ═══════════════════════════════════════════════════════════════
import hashlib
import json
import os
import time
import urllib.error
import urllib.request
import urllib.parse
from datetime import date

try:
    from dotenv import load_dotenv
    load_dotenv(SCRIPT_DIR / ".env")
except ImportError:
    pass  # dotenv не обязателен — можно просто выставить переменную окружения

DEEPL_API_KEY = os.environ.get("DEEPL_API_KEY", "").strip()
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "").strip()
ARBITER_MODEL = "claude-haiku-4-5-20251001"  # GUARDRAILS #2: не менять без просьбы Ra
TRANSLATIONS_JSON_PATH = SCRIPT_DIR / "translations.json"
REPORTS_DIR = SCRIPT_DIR / "reports"

def _deepl_base_url(api_key):
    # Free-тариф — ключи оканчиваются на ":fx", у него отдельный домен API.
    return "https://api-free.deepl.com/v2/translate" if api_key.endswith(":fx") \
        else "https://api.deepl.com/v2/translate"

def _deepl_request(params, api_key):
    """POST к DeepL. Ретрай до 5 попыток только на 429 (rate limit), растущая
    пауза с потолком 30с — прочие ошибки сразу наверх, глушить их нельзя.
    5, а не 3: этап G шлёт altEn по одному слову (308 запросов подряд) и на
    3 попытках ловил 429; лишняя терпеливость дешевле флага-ложняка."""
    data = urllib.parse.urlencode(params).encode("utf-8")
    req = urllib.request.Request(_deepl_base_url(api_key), data=data, headers={
        "Authorization": f"DeepL-Auth-Key {api_key}",
        "Content-Type": "application/x-www-form-urlencoded",
    })
    last_err = None
    for attempt in range(5):
        try:
            with urllib.request.urlopen(req, timeout=30) as resp:
                return json.loads(resp.read().decode("utf-8"))
        except urllib.error.HTTPError as e:
            last_err = e
            if e.code == 429 and attempt < 4:
                time.sleep(min(2 ** attempt, 30))
                continue
            raise
    raise last_err

def _deepl_translate_batch(texts, api_key, source_lang="RU", target_lang="EN"):
    """Один запрос к DeepL на пачку строк. Возвращает переводы в том же порядке."""
    params = [("text", t) for t in texts] + [("source_lang", source_lang), ("target_lang", target_lang)]
    result = _deepl_request(params, api_key)
    return [tr["text"] for tr in result["translations"]]

def _deepl_translate_one(text, api_key, source_lang, target_lang, context=None):
    """Перевод одной строки с опциональным контекстом. Контекст не тарифицируется,
    но DeepL считает его один на весь запрос — а у каждого VOCAB-слова свой
    (артикль/часть речи/пример), поэтому тут не пачка, а один текст за раз."""
    params = [("text", text), ("source_lang", source_lang), ("target_lang", target_lang)]
    if context:
        params.append(("context", context))
    result = _deepl_request(params, api_key)
    return result["translations"][0]["text"]

# ─── translations.json — кэш и источник правды для VOCAB.en/RULES/TERMS/SOUNDS
# (см. BRIEF §3). Ведёт Claude Code, Ra файл не открывает — поэтому пустой
# или битый файл это ошибка, а не сигнал молча перевести всё заново.
TRANSLATIONS_SECTIONS = ("VOCAB", "CONJUGATIONS", "REGEL_VERBS", "RULES", "TERMS", "SOUNDS")

def load_translations(warn):
    """Единственный кэш и источник правды для всех переводимых полей.
    Старого механизма (диф по id против предыдущего data.js регуляркой,
    _parse_old_array_by_id) больше нет — удалён на этапе G: id перевыдаётся
    каждый прогон, кэш промахивался блоками, при неудачном парсинге функция
    молча возвращала пустоту и весь словарь уходил в переперевод."""
    empty = {k: {} for k in TRANSLATIONS_SECTIONS}
    if not TRANSLATIONS_JSON_PATH.exists():
        warn.append("translations.json не найден — будет создан заново, первый прогон переведёт всё")
        return empty
    try:
        data = json.loads(TRANSLATIONS_JSON_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as e:
        warn.append(f"translations.json не читается ({e}) — переводы в этом прогоне НЕ ОБНОВЛЯЮТСЯ "
                    f"(ни новые, ни повторное использование кэша), чтобы не потерять то, что уже "
                    f"вручную починено. Почини файл (см. git-историю) и прогони заново.")
        return None
    for k in TRANSLATIONS_SECTIONS:
        data.setdefault(k, {})
    return data

def save_translations(data):
    TRANSLATIONS_JSON_PATH.write_text(
        json.dumps(data, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")

def simple_hash(*parts):
    raw = "|".join(p or "" for p in parts)
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:16]

def vocab_hash(item):
    """Ключ VOCAB в translations.json — хэш источника (de+gender+pos+exampleDe),
    не id: assign_ids перевыдаёт id при каждом прогоне, ключ по id промахивается
    блоками, стоит добавить/удалить слово выше по списку. exampleDe — часть
    хэша не случайно: он же часть контекста для перевода en (см. ниже), так
    что смена примера обоснованно инвалидирует и слово тоже."""
    return simple_hash(item.get("de"), item.get("gender"), item.get("pos"), item.get("exampleDe"))

POS_DE_LABEL = {"noun": "Nomen", "verb": "Verb", "adj": "Adjektiv", "adv": "Adverb",
                "pron": "Pronomen", "num": "Numerale", "phrase": "Wendung"}
GENDER_ARTICLE = {"m": "der", "f": "die", "n": "das", "pl": "die"}

def build_vocab_context(item):
    """Контекст для DeepL — не переводится и не тарифицируется, только
    подсказка модели: артикль, часть речи, пример, альт. форма."""
    parts = []
    article = GENDER_ARTICLE.get(item.get("gender"))
    if article:
        parts.append(article)
    pos_label = POS_DE_LABEL.get(item.get("pos"))
    if pos_label:
        parts.append(pos_label)
    if item.get("exampleDe"):
        parts.append(item["exampleDe"])
    if item.get("altDe"):
        parts.append(item["altDe"])
    return " | ".join(parts) if parts else None

def normalize_verb_en(text):
    """«to <verb>» — одна форма на весь словарь, независимо от того, что
    вернул DeepL (с to / без / Capitalized)."""
    text = (text or "").strip()
    if not text:
        return text
    m = re.match(r"^to\s+(.*)$", text, re.IGNORECASE)
    core = m.group(1) if m else text
    return "to " + core

# Закрытый список — единственная категория собственных имён, реально
# встречающаяся в VOCAB (проверено: ни одной отдельной записи для стран/
# языков/личных имён в словаре нет, только дни недели и месяцы).
PROPER_NOUN_EN = {
    "monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday",
    "january", "february", "march", "april", "may", "june", "july",
    "august", "september", "october", "november", "december",
}

def is_proper_noun_en(text):
    return (text or "").strip().lower() in PROPER_NOUN_EN

def normalize_vocab_en(text, pos):
    """pos==verb → normalize_verb_en(); verb/adj/adv/noun → строчная первая
    буква — по-английски заглавная у DeepL для существительных унаследована
    из немецкого написания, а не грамматика. Исключение — закрытый список
    настоящих собственных (дни недели, месяцы)."""
    text = (text or "").strip()
    if not text:
        return text
    if pos == "verb":
        text = normalize_verb_en(text)
    if pos in ("verb", "adj", "adv") and text:
        text = text[0].lower() + text[1:]
    elif pos == "noun" and text and not is_proper_noun_en(text):
        text = text[0].lower() + text[1:]
    return text

def build_backtranslation_text(item, en_text):
    """Голый EN-текст неоднозначен для DeepL без артикля/инфинитив-маркера —
    он читает его как что придётся (для глаголов — как цель: "to work" →
    "um zu arbeiten"/"zur Arbeit" вместо инфинитива; для существительных
    без артикля — не факт что как существительное вообще). Возвращаем
    артикль/маркер перед обратным переводом, снимаем его сравнением через
    normalize_de_for_compare (уже снимает der/die/das)."""
    pos = item.get("pos")
    if pos == "verb":
        m = re.match(r"^to\s+(.*)$", en_text, re.IGNORECASE)
        return m.group(1) if m else en_text
    if pos == "noun" and not is_proper_noun_en(en_text):
        return "the " + en_text
    return en_text

def has_cyrillic(s):
    return bool(re.search(r"[А-Яа-яЁё]", s or ""))

def normalize_de_for_compare(s):
    """Для сверки обратного перевода: нижний регистр, без артикля."""
    s = (s or "").strip().lower()
    return re.sub(r"^(der|die|das)\s+", "", s)

def compute_vocab_plan(all_vocab, translations, warn):
    """Без обращения к API: делит all_vocab на кэш-хиты (уже есть en в
    translations.json под текущим хэшем источника) и кандидатов на перевод.
    Считает символы прямого перевода (de) для --dry-run. Если translations.json
    не читается (load_translations вернул None) — кандидатов нет вообще: это
    единственный источник кэша, фолбэка на старый data.js больше нет."""
    candidates = []
    cache_hits = 0
    if translations is None:
        return candidates, cache_hits, 0
    vocab_cache = translations["VOCAB"]
    for item in all_vocab:
        if not item.get("de"):
            continue
        h = vocab_hash(item)
        cached = vocab_cache.get(h)
        if item.get("new") and cached:
            warn.append(f"VOCAB: «{item['de']}» помечено new=TRUE в xlsx, но перевод уже есть "
                        f"в translations.json (hash {h}) — сверь вручную, скрипт не решает сам")
        if cached and cached.get("en"):
            cache_hits += 1
            en_val = cached["en"]
            # Самоисправление: если нормализация поменялась (напр. строчная
            # буква у существительных), применяем её и к уже закэшированным
            # deepl-переводам — без обращения к API, просто локальная строка.
            # source: "manual" не трогаем никогда (§26 GUARDRAILS).
            if cached.get("source") == "deepl":
                renorm = normalize_vocab_en(en_val, item.get("pos"))
                if renorm != en_val:
                    cached["en"] = renorm
                    en_val = renorm
            item["en"] = en_val
        else:
            candidates.append(item)
    char_count = sum(len(it["de"]) for it in candidates)
    return candidates, cache_hits, char_count

def check_vocab_threshold(n_candidates, all_vocab, force):
    """True — можно продолжать. False — превышен порог 20%, нужен --force."""
    total = sum(1 for it in all_vocab if it.get("de"))
    if total == 0:
        return True
    ratio = n_candidates / total
    if ratio > 0.20 and not force:
        print(f"\n✗ ОСТАНОВЛЕНО: перевод затронет {n_candidates}/{total} слов VOCAB "
              f"({ratio:.0%}) — больше порога 20%.")
        print("  Если это ожидаемо (первый прогон, массовая правка xlsx) — повтори с --force.")
        return False
    return True

def run_vocab_translation(candidates, vocab_cache, warn):
    """Реальный прогон: DE→EN + контекст (фаза 1), обратный перевод EN→DE
    (фаза 2), сверка. item['en'] пишется только для прошедших обе проверки;
    остальные уходят в report_rows (поле en НЕ пишется — сайт покажет ru,
    это штатный исход, не ошибка). Кэш обновляется только для успешных —
    флаги не кэшируются, чтобы не запирать неразрешённый случай навсегда.

    Фаза 1 — по одному слову за запрос: у DeepL context один на весь
    запрос, а не на текст внутри пачки, а у каждого VOCAB-слова свой
    контекст (артикль/часть речи/пример) — batch тут физически невозможен
    без потери контекста, ради которого его и завели.
    Фаза 2 — контекста не требует (обратный перевод переводит голое слово),
    поэтому batch-ится обычным способом, BATCH=50."""
    if not DEEPL_API_KEY:
        warn.append("DEEPL_API_KEY не задан — VOCAB.en не переведён")
        return [], 0
    today = date.today().isoformat()
    BATCH = 50

    # Фаза 1: DE→EN + контекст, по одному слову — см. докстринг.
    print(f"  → VOCAB.en фаза 1/2 (DE→EN + контекст, по слову): {len(candidates)}...")
    forward_ok = []  # (item, en_text)
    report_rows = []
    for i, item in enumerate(candidates, 1):
        context = build_vocab_context(item)
        try:
            raw_en = _deepl_translate_one(item["de"], DEEPL_API_KEY, "DE", "EN", context=context)
        except Exception as e:
            warn.append(f"DeepL: ошибка перевода VOCAB «{item['de']}»: {e}")
            report_rows.append({"item": item, "en_deepl": "(ошибка API)", "back_de": "—"})
            continue
        en_text = normalize_vocab_en(raw_en, item.get("pos"))
        if has_cyrillic(en_text):
            report_rows.append({"item": item, "en_deepl": en_text, "back_de": "(кириллица в результате)"})
            continue
        forward_ok.append((item, en_text))
        if i % 100 == 0:
            print(f"    ... {i}/{len(candidates)}")

    # Фаза 2: обратный перевод EN→DE, пачками — без контекста, batch допустим.
    print(f"  → VOCAB.en фаза 2/2 (обратный перевод EN→DE, пачками по {BATCH}): "
          f"{len(forward_ok)}...")
    ok = 0
    for i in range(0, len(forward_ok), BATCH):
        chunk = forward_ok[i:i + BATCH]
        bt_texts = [build_backtranslation_text(item, en_text) for item, en_text in chunk]
        try:
            back_translations = _deepl_translate_batch(
                bt_texts, DEEPL_API_KEY, source_lang="EN", target_lang="DE")
        except Exception as e:
            warn.append(f"DeepL: ошибка обратного перевода батч {i}-{i + len(chunk)}: {e}")
            for item, en_text in chunk:
                report_rows.append({"item": item, "en_deepl": en_text, "back_de": "(ошибка API)"})
            continue
        for (item, en_text), back_de in zip(chunk, back_translations):
            if normalize_de_for_compare(back_de) != normalize_de_for_compare(item["de"]):
                report_rows.append({"item": item, "en_deepl": en_text, "back_de": back_de})
                continue
            item["en"] = en_text
            vocab_cache[vocab_hash(item)] = {"de": item["de"], "en": en_text, "source": "deepl",
                                              "backtranslation": back_de, "date": today}
            ok += 1
        print(f"    ... {min(i + BATCH, len(forward_ok))}/{len(forward_ok)}")
    print(f"  ✓ VOCAB.en: переведено и подтверждено {ok}/{len(candidates)}, "
          f"флагов на разбор: {len(report_rows)}")
    return report_rows, ok

def write_vocab_review_report(report_rows):
    REPORTS_DIR.mkdir(exist_ok=True)
    path = REPORTS_DIR / "en_review_vocab.md"
    if not report_rows:
        path.write_text(
            "# en_review_vocab — флагов нет\n\n"
            "Все переводы этого прогона прошли обратный перевод и проверку на кириллицу.\n",
            encoding="utf-8")
        return path
    rows = sorted(report_rows, key=lambda r: (r["item"].get("pos") or "", r["item"].get("level") or "",
                                               r["item"].get("de") or ""))
    lines = [
        "# en_review_vocab — флаги обратного перевода",
        "",
        f"Прогон: {date.today().isoformat()}. Всего флагов: {len(rows)}.",
        "Слово НЕ попало в data.js (поле `en` не записано, сайт покажет русский), "
        "пока не разберёшь тут и не занесёшь решение в translations.json.",
        "",
        "| id | de | ru | en (DeepL) | обратно в DE | pos | level |",
        "|---|---|---|---|---|---|---|",
    ]
    for r in rows:
        it = r["item"]
        lines.append(f"| {it.get('id','')} | {it.get('de','')} | {it.get('ru','') or ''} | "
                      f"{r.get('en_deepl','')} | {r.get('back_de','')} | {it.get('pos','')} | "
                      f"{it.get('level','') or ''} |")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return path

# ═══════════════════════════════════════════════════════════════
# АРБИТР НА HAIKU (--arbiter=haiku, режим миграции, BRIEF §7)
# Разбирает ТОЛЬКО флаги обратного перевода VOCAB.en — не весь словарь,
# не RULES/TERMS/SOUNDS (GUARDRAILS #28, #32). source: "haiku", не
# "manual" (#27) — смешивать машинный разбор с человеческим запрещено.
# ═══════════════════════════════════════════════════════════════
# Строгая схема вердикта — structured output через tool use, а не текстовый
# JSON: модель физически не может обернуть ответ в ```, потому что ответ
# приходит не текстом, а готовым разобранным полем tool_use.input. Это
# решает задачу на уровне протокола (см. обсуждение этапа B — попытка
# "распарсить ```json...```" через строгий json.loads() на голом тексте
# дала 100% skip: Haiku оборачивает JSON в markdown-фенсы вопреки прямому
# запрету в промпте, и это не разово — второй артефакт после "to "/"the ").
ARBITER_TOOL = {
    "name": "submit_verdicts",
    "description": "Вердикты по батчу слов словаря: ok / fix / skip для каждого id из запроса.",
    "input_schema": {
        "type": "object",
        "properties": {
            "verdicts": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "id": {"type": "string"},
                        "verdict": {"type": "string", "enum": ["ok", "fix", "skip"]},
                        "en": {"type": "string"},
                        "reason": {"type": "string"},
                    },
                    "required": ["id", "verdict"],
                    "additionalProperties": False,
                },
            },
        },
        "required": ["verdicts"],
        "additionalProperties": False,
    },
}

def _anthropic_request(system_prompt, user_content, api_key, max_tokens=4096):
    """POST к Anthropic Messages API с принудительным tool use (см. ARBITER_TOOL
    выше). Ретрай до 3 попыток только на 429, остальные ошибки сразу наверх —
    та же дисциплина, что у _deepl_request."""
    payload = json.dumps({
        "model": ARBITER_MODEL,
        "max_tokens": max_tokens,
        "system": system_prompt,
        "messages": [{"role": "user", "content": user_content}],
        "tools": [ARBITER_TOOL],
        "tool_choice": {"type": "tool", "name": "submit_verdicts"},
    }).encode("utf-8")
    req = urllib.request.Request(
        "https://api.anthropic.com/v1/messages",
        data=payload,
        headers={
            "x-api-key": api_key,
            "anthropic-version": "2023-06-01",
            "content-type": "application/json",
        },
    )
    last_err = None
    for attempt in range(3):
        try:
            with urllib.request.urlopen(req, timeout=60) as resp:
                return json.loads(resp.read().decode("utf-8"))
        except urllib.error.HTTPError as e:
            last_err = e
            if e.code == 429 and attempt < 2:
                time.sleep(2 * (attempt + 1))
                continue
            raise
    raise last_err

ARBITER_SYSTEM_PROMPT = """Ты — арбитр в пайплайне перевода немецко-русского словаря на английский (DE→EN).

Каждое слово уже прошло: 1) перевод DeepL DE→EN (поле en_deepl), 2) обратный перевод EN→DE для проверки (поле back_de). Обратный перевод разошёлся с оригиналом (de) — это может быть ложное срабатывание (en_deepl верен, просто обратный перевод выбрал синоним или другую форму), а может быть реальная ошибка перевода.

Для каждого слова в батче верни вердикт:
- "ok" — en_deepl верен, расхождение — ложное срабатывание обратного перевода.
- "fix" — en_deepl неверен, в поле "en" укажи правильный перевод.
- "skip" — не уверен (многозначное слово, недостаточно контекста).

Слова — из учебного словаря уровня A1–A2 (поле "level" в запросе). Выбирай
перевод, который реально преподают на этом уровне, а не редкий/книжный
синоним, даже если он тоже формально верен: "cheap", не "inexpensive";
"begin", не "commence".

При выборе значения многозначного слова опирайся на exampleDe и note —
это конкретный контекст, в котором слово используется в этом словаре, а
не на самое частотное значение слова в языке вообще.

Конвенции для поля "en", если verdict=fix:
- ровно один вариант перевода, без слэшей и без "/" (не "cheap/inexpensive").
- глаголы — форма "to <verb>", один глагол.
- прилагательные/наречия — со строчной буквы.
- существительные — строчная буква, кроме настоящих собственных (дни недели, месяцы).

Вызови submit_verdicts ровно один раз, с одним элементом verdicts на каждый id из входного массива (в любом порядке). Для verdict="fix" заполни поле "en". Для verdict="skip" заполни поле "reason" (коротко, по-русски)."""

def build_arbiter_batch_payload(rows):
    items = []
    for r in rows:
        it = r["item"]
        items.append({
            "id": it.get("id"),
            "de": it.get("de"),
            "article": GENDER_ARTICLE.get(it.get("gender")),
            "pos": it.get("pos"),
            "level": it.get("level"),
            "ru": it.get("ru"),
            "altRu": it.get("altRu"),
            "exampleDe": it.get("exampleDe"),
            "note": it.get("note"),
            "en_deepl": r.get("en_deepl"),
            "back_de": r.get("back_de"),
        })
    return json.dumps(items, ensure_ascii=False)

def run_haiku_arbiter(report_rows, vocab_cache, warn):
    """Возвращает (remaining_rows, stats). remaining_rows — то, что осталось
    неразобранным (skip + сбои парсинга) и идёт в reports/en_review_vocab.md
    как обычно. Разобранное (ok/fix) убирается из отчёта на ручной разбор —
    оно уже разобрано, просто не человеком."""
    empty_stats = {"ok": 0, "fix": 0, "skip": 0, "fix_list": [], "skip_list": [],
                   "input_tokens": 0, "output_tokens": 0}
    if not report_rows:
        return report_rows, empty_stats
    if not ANTHROPIC_API_KEY:
        warn.append("ANTHROPIC_API_KEY не задан — арбитр Haiku пропущен, все флаги остаются на ручной разбор")
        return report_rows, empty_stats

    BATCH = 40
    today = date.today().isoformat()
    stats = dict(empty_stats)
    stats["fix_list"] = []
    stats["skip_list"] = []
    remaining_rows = []

    print(f"  → Арбитр Haiku ({ARBITER_MODEL}): {len(report_rows)} флагов, батчи по {BATCH}...")
    for i in range(0, len(report_rows), BATCH):
        chunk = report_rows[i:i + BATCH]
        chunk_by_id = {r["item"].get("id"): r for r in chunk}
        try:
            resp = _anthropic_request(ARBITER_SYSTEM_PROMPT, build_arbiter_batch_payload(chunk), ANTHROPIC_API_KEY)
        except Exception as e:
            warn.append(f"Arbiter: ошибка запроса батч {i}-{i + len(chunk)}: {e}")
            for r in chunk:
                it = r["item"]
                stats["skip"] += 1
                stats["skip_list"].append({"id": it.get("id"), "de": it.get("de"), "reason": f"ошибка API: {e}"})
                remaining_rows.append(r)
            continue

        usage = resp.get("usage", {})
        stats["input_tokens"] += usage.get("input_tokens", 0)
        stats["output_tokens"] += usage.get("output_tokens", 0)

        # Structured output (tool use) — ответ уже разобранный JSON в
        # tool_use.input, не текст. Не пришёл tool_use с ожидаемой формой —
        # весь батч в skip (та же дисциплина #30, но на уровне протокола).
        tool_block = next((b for b in resp.get("content", [])
                            if b.get("type") == "tool_use" and b.get("name") == "submit_verdicts"), None)
        verdicts = (tool_block or {}).get("input", {}).get("verdicts") if tool_block else None
        if not isinstance(verdicts, list):
            warn.append(f"Arbiter: нет tool_use submit_verdicts в ответе, батч {i}-{i + len(chunk)} — весь батч в skip")
            for r in chunk:
                it = r["item"]
                stats["skip"] += 1
                stats["skip_list"].append({"id": it.get("id"), "de": it.get("de"),
                                            "reason": "ответ модели не распарсился (весь батч)"})
                remaining_rows.append(r)
            continue

        verdicts_by_id = {v["id"]: v for v in verdicts if isinstance(v, dict) and v.get("id")}
        for req_id, r in chunk_by_id.items():
            it = r["item"]
            v = verdicts_by_id.get(req_id)
            # id из ответа сверяется с id из запроса — не совпал (или нет ответа) → skip (#25 брифа).
            if not v:
                stats["skip"] += 1
                stats["skip_list"].append({"id": req_id, "de": it.get("de"), "reason": "нет ответа модели по этому id"})
                remaining_rows.append(r)
                continue
            verdict = v.get("verdict")
            if verdict == "ok":
                en_text = r.get("en_deepl")
                vocab_cache[vocab_hash(it)] = {"de": it.get("de"), "en": en_text, "source": "haiku",
                                                "verdict": "ok", "backtranslation": r.get("back_de"), "date": today}
                it["en"] = en_text
                stats["ok"] += 1
            elif verdict == "fix":
                en_raw = v.get("en")
                if not en_raw:
                    stats["skip"] += 1
                    stats["skip_list"].append({"id": req_id, "de": it.get("de"), "reason": "verdict=fix без поля en"})
                    remaining_rows.append(r)
                    continue
                en_text = normalize_vocab_en(en_raw, it.get("pos"))
                vocab_cache[vocab_hash(it)] = {"de": it.get("de"), "en": en_text, "source": "haiku",
                                                "verdict": "fix", "date": today}
                it["en"] = en_text
                stats["fix"] += 1
                stats["fix_list"].append({"id": req_id, "de": it.get("de"), "was": r.get("en_deepl"), "now": en_text})
            elif verdict == "skip":
                stats["skip"] += 1
                stats["skip_list"].append({"id": req_id, "de": it.get("de"), "reason": v.get("reason") or "модель не уверена"})
                remaining_rows.append(r)
            else:
                stats["skip"] += 1
                stats["skip_list"].append({"id": req_id, "de": it.get("de"), "reason": f"неизвестный verdict: {verdict!r}"})
                remaining_rows.append(r)
        print(f"    ... {min(i + BATCH, len(report_rows))}/{len(report_rows)}")

    print(f"  ✓ Арбитр: ok={stats['ok']} fix={stats['fix']} skip={stats['skip']}, "
          f"токены Anthropic: {stats['input_tokens']} вход / {stats['output_tokens']} выход")
    return remaining_rows, stats

def write_arbiter_report(stats):
    REPORTS_DIR.mkdir(exist_ok=True)
    path = REPORTS_DIR / "en_arbiter.md"
    total = stats["ok"] + stats["fix"] + stats["skip"]
    lines = [
        "# en_arbiter — разбор флагов через Claude Haiku",
        "",
        f"Прогон: {date.today().isoformat()}. Модель: {ARBITER_MODEL}.",
        f"Всего флагов на входе: {total}. ok: {stats['ok']} · fix: {stats['fix']} · skip: {stats['skip']}.",
        f"Токены Anthropic: {stats['input_tokens']} вход / {stats['output_tokens']} выход.",
        "",
        "## fix — кандидат DeepL был неверен, модель дала правильный вариант",
        "",
    ]
    if stats["fix_list"]:
        lines += ["| id | de | было (DeepL) | стало (Haiku) |", "|---|---|---|---|"]
        for f in sorted(stats["fix_list"], key=lambda x: x["de"] or ""):
            lines.append(f"| {f['id']} | {f['de']} | {f['was']} | {f['now']} |")
    else:
        lines.append("(пусто)")
    lines += ["", "## skip — модель не уверена, остаются на разбор в чате (см. en_review_vocab.md)", ""]
    if stats["skip_list"]:
        lines += ["| id | de | причина |", "|---|---|---|"]
        for s in sorted(stats["skip_list"], key=lambda x: x["de"] or ""):
            lines.append(f"| {s['id']} | {s['de']} | {s['reason']} |")
    else:
        lines.append("(пусто)")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return path

# ═══════════════════════════════════════════════════════════════
# RULES (этап C, BRIEF §C) — гибрид: 31 исключение из docs/en-fix/
# translations-rules-batch*.json (source: manual, никогда не пересчитываются),
# остальные 28 — DeepL RU→EN с ignore-тегами вокруг латинских (=немецких)
# фрагментов. Кэш — translations.json["RULES"] по id (не хэш: id у RULES
# стабильны, правятся редко — см. BRIEF §3). Диф — свой, без старого
# data.js: каждая запись кэша хранит _src_<field> (значение поля-источника
# на момент перевода) — иначе старый RU-пивот перевод считался бы «уже
# переведённым» навсегда, раз поле-источник не менялось. Тот же приём —
# _src_altEn на этапе G.
# ═══════════════════════════════════════════════════════════════
LATIN_WORD_RE = re.compile(r"[A-Za-zÀ-ÖØ-öø-ÿ][A-Za-zÀ-ÖØ-öø-ÿ'\-]*")

def wrap_latin_ignore_tags(text):
    """Оборачивает каждый латинский токен в <x>…</x>. content_md/title/note
    у RULES — русская проза кириллицей, поэтому любой латинский фрагмент по
    определению немецкий (см. BRIEF §C). Markdown (**, |, →, 💡, ⚠) не входит
    в класс символов регэкспа — не заворачивается, `**muss**` → `**<x>muss</x>**`."""
    return LATIN_WORD_RE.sub(lambda m: f"<x>{m.group(0)}</x>", text)

def strip_ignore_tags(text):
    return text.replace("<x>", "").replace("</x>", "")

def check_rule_field_invariants(original, translated):
    """(ok, reason). Все четыре инварианта из BRIEF §C: кириллицы нет,
    число строк совпадает, число | в каждой строке совпадает, каждый
    латинский токен исходника присутствует в переводе дословно."""
    if has_cyrillic(translated):
        return False, "кириллица в результате"
    orig_lines = original.split("\n")
    tr_lines = translated.split("\n")
    if len(orig_lines) != len(tr_lines):
        return False, f"число строк не совпадает ({len(orig_lines)} → {len(tr_lines)})"
    for idx, (ol, tl) in enumerate(zip(orig_lines, tr_lines), 1):
        if ol.count("|") != tl.count("|"):
            return False, f"строка {idx}: число | не совпадает ({ol.count('|')} → {tl.count('|')})"
    missing = [tok for tok in dict.fromkeys(LATIN_WORD_RE.findall(original)) if tok not in translated]
    if missing:
        return False, f"пропали токены: {', '.join(missing[:6])}"
    return True, None

def _deepl_translate_batch_xml(texts, api_key, source_lang="RU", target_lang="EN", ignore_tags="x"):
    params = [("text", t) for t in texts] + [
        ("source_lang", source_lang), ("target_lang", target_lang),
        ("tag_handling", "xml"), ("ignore_tags", ignore_tags),
    ]
    result = _deepl_request(params, api_key)
    return [tr["text"] for tr in result["translations"]]

def load_manual_rules_batches(translations, warn):
    """Подтягивает docs/en-fix/translations-rules-batch*.json в
    translations.json['RULES'] с source: manual. Источник — ручная работа,
    авторитетный: перезаписывает существующую запись того же id (в отличие
    от DeepL-переводов, которые никогда не трогают manual)."""
    if translations is None:
        return []
    cache = translations["RULES"]
    batch_dir = SCRIPT_DIR / "docs" / "en-fix"
    files = sorted(batch_dir.glob("translations-rules-batch*.json")) if batch_dir.exists() else []
    loaded_ids = []
    for f in files:
        try:
            data = json.loads(f.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError) as e:
            warn.append(f"RULES manual batch {f.name}: не читается ({e})")
            continue
        for rid, entry in data.get("RULES", {}).items():
            if entry.get("source") != "manual":
                warn.append(f"RULES manual batch {f.name}: запись {rid} без source=manual — пропущена")
                continue
            cache[rid] = dict(entry)
            loaded_ids.append(rid)
    if loaded_ids:
        print(f"  ✓ RULES manual: {len(loaded_ids)} записей из {len(files)} batch-файлов")
    return loaded_ids

def apply_manual_rules(rules, translations, warn):
    """source: manual — всегда из кэша, никогда не пересчитывается (#26)."""
    manual_ids = set()
    if translations is None:
        return manual_ids
    cache = translations["RULES"]
    for r in rules:
        cached = cache.get(r.get("id"))
        if cached and cached.get("source") == "manual":
            for field, en_field in (("title", "titleEn"), ("content_md", "content_md_en"),
                                     ("note", "noteEn"), ("examples", "examplesEn")):
                if cached.get(en_field):
                    r[en_field] = cached[en_field]
            manual_ids.add(r["id"])
    return manual_ids

def compute_ignoretag_candidates(items, field, en_field, cache, key_fn, manual_ids=frozenset()):
    """Обобщённая версия — этап C (RULES) и этап D (VOCAB.note/warning,
    TERMS.note). key_fn достаёт ключ кэша (id для RULES/TERMS, vocab_hash
    для VOCAB — у VOCAB кэш общий с основным словом, тот же хэш)."""
    candidates = []
    if cache is None:
        return candidates
    for it in items:
        key = key_fn(it)
        if key in manual_ids:
            continue
        val = it.get(field)
        if not val:
            continue
        cached = cache.get(key) or {}
        if cached.get("source") == "manual":
            # #26: manual не пересчитывается. Но уже переведённый en_field
            # применяем как есть — ручной noteEn из translations-vocab-skip.json
            # либо машинный, попавший в ту же ячейку кэша, что и ручной en.
            # Без этого загрузка skip-слова тихо снимала бы у слова английское
            # примечание (source: manual выключал этап D для всей ячейки).
            # Как и manual en в compute_vocab_plan — применяется безусловно,
            # свежесть manual-значений на человеке.
            if cached.get(en_field):
                it[en_field] = cached[en_field]
            continue
        if cached.get(en_field) and cached.get(f"_src_{field}") == val:
            it[en_field] = cached[en_field]
        else:
            candidates.append(it)
    return candidates

def translate_ignoretag_field(candidates, field, en_field, cache, key_fn, warn, label, batch_size=40):
    """RU→EN с ignore-тегами, ПОСТРОЧНО — не всем полем целиком одним
    DeepL-запросом. Эмпирически найдено при первом прогоне: если в
    многострочном тексте на одной строке несколько <x>-тегов, DeepL's
    tag_handling=xml непредсказуемо расставляет переводы разных строк по
    отдельным строкам вывода (число строк расходилось в разы, 8→31 и
    хуже) — обёртка ломала структуру ровно так, как предупреждал стоп-
    критерий брифа. non_splitting_tags/outline_detection=0 не помогли
    (первый вообще снял защиту ignore_tags). Однострочный вход — всегда
    ровно однострочный выход, без исключений (проверено отдельно) —
    поэтому переводим по одной строке за раз, объединяя строки разных
    правил в общий батч для API. Поле целиком немецкое (кириллицы нет —
    22 таких title) копируется как есть, без API: переводить нечего.
    Инварианты — на каждую запись отдельно; не прошла → en_field не
    пишется, строка в отчёт, штатный откат на русский (не ошибка)."""
    report_rows = []
    if not candidates:
        return report_rows
    today = date.today().isoformat()
    verbatim = [it for it in candidates if not has_cyrillic(it[field])]
    need_api = [it for it in candidates if has_cyrillic(it[field])]
    for it in verbatim:
        val = it[field]
        it[en_field] = val
        entry = cache.setdefault(key_fn(it), {})
        entry.setdefault("source", "deepl")
        entry[f"_src_{field}"] = val
        entry[en_field] = val
        entry["date"] = today
    if not need_api:
        return report_rows
    if not DEEPL_API_KEY:
        warn.append(f"DEEPL_API_KEY не задан — {label}.{en_field} не переведён ({len(need_api)} записей)")
        return report_rows
    print(f"  → {label}.{field} → {en_field} (ignore-теги, RU→EN, построчно): {len(need_api)}...")

    # Разворачиваем все записи в плоский список строк (границы — по item_lines).
    item_lines = [it[field].split("\n") for it in need_api]
    flat_lines = [line for lines in item_lines for line in lines]
    translated_flat = [None] * len(flat_lines)
    to_send = [(i, wrap_latin_ignore_tags(line)) for i, line in enumerate(flat_lines) if line.strip()]
    for i, line in enumerate(flat_lines):
        if not line.strip():
            translated_flat[i] = line  # пустая строка — нечего слать в API

    for i in range(0, len(to_send), batch_size):
        chunk = to_send[i:i + batch_size]
        try:
            out = _deepl_translate_batch_xml([w for _idx, w in chunk], DEEPL_API_KEY)
        except Exception as e:
            warn.append(f"DeepL: ошибка перевода {label}.{field} батч строк {i}-{i + len(chunk)}: {e}")
            continue  # translated_flat[idx] остаётся None → эта запись уйдёт в отчёт ниже
        for (idx, _w), tr in zip(chunk, out):
            translated_flat[idx] = strip_ignore_tags(tr)

    ok = 0
    pos = 0
    for it, lines in zip(need_api, item_lines):
        n = len(lines)
        chunk_tr = translated_flat[pos:pos + n]
        pos += n
        key = key_fn(it)
        name = it.get("de") or it.get("term") or it.get("title") or it.get("id") or ""
        if any(t is None for t in chunk_tr):
            report_rows.append({"id": key, "label": label, "name": name, "field": field,
                                 "reason": "ошибка API (см. WARN)"})
            continue
        tr_joined = "\n".join(chunk_tr)
        is_ok, reason = check_rule_field_invariants(it[field], tr_joined)
        if not is_ok:
            report_rows.append({"id": key, "label": label, "name": name, "field": field, "reason": reason})
            continue
        it[en_field] = tr_joined
        entry = cache.setdefault(key, {})
        entry.setdefault("source", "deepl")
        entry[f"_src_{field}"] = it[field]
        entry[en_field] = tr_joined
        entry["date"] = today
        ok += 1
    print(f"  ✓ {label}.{en_field}: переведено {ok}/{len(need_api)}, не прошло инвариант/ошибка: {len(need_api) - ok}")
    return report_rows

def write_rules_review_report(rules, translations, field_report_rows):
    """По каждому из 59 правил: source, что переведено, что не прошло."""
    REPORTS_DIR.mkdir(exist_ok=True)
    path = REPORTS_DIR / "en_review_rules.md"
    cache = translations["RULES"] if translations else {}
    fails_by_id = {}
    for row in field_report_rows:
        fails_by_id.setdefault(row["id"], []).append(f"{row['field']}: {row['reason']}")
    lines = [
        "# en_review_rules — статус перевода правил",
        "",
        f"Прогон: {date.today().isoformat()}. Всего правил: {len(rules)}.",
        "",
        "| id | title | source | titleEn | content_md_en | noteEn | не прошло инварианты |",
        "|---|---|---|---|---|---|---|",
    ]
    for r in sorted(rules, key=lambda x: x["id"]):
        rid = r["id"]
        cached = cache.get(rid, {})
        source = cached.get("source", "—")
        has_title = "ok" if r.get("titleEn") else "—"
        has_content = "ok" if r.get("content_md_en") else "—"
        has_note = "ok" if (not r.get("note")) or r.get("noteEn") else "—"
        fails = "; ".join(fails_by_id.get(rid, [])) or ""
        title_short = (r.get("title") or "")[:45]
        lines.append(f"| {rid} | {title_short} | {source} | {has_title} | {has_content} | {has_note} | {fails} |")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return path

def write_notes_review_report(report_rows):
    """Этап D: флаги ignore-тегов для VOCAB.note/warning и TERMS.note —
    не прошедшие инварианты, поле не записано, откат на русский."""
    REPORTS_DIR.mkdir(exist_ok=True)
    path = REPORTS_DIR / "en_review_notes.md"
    if not report_rows:
        path.write_text("# en_review_notes — флагов нет\n\nВсе переводы прошли инварианты.\n", encoding="utf-8")
        return path
    rows = sorted(report_rows, key=lambda r: (r["label"], r["name"] or ""))
    lines = [
        "# en_review_notes — флаги ignore-тегов (VOCAB.note/warning, TERMS.note)",
        "",
        f"Прогон: {date.today().isoformat()}. Всего флагов: {len(rows)}.",
        "Поле не попало в data.js (сайт покажет русский), пока не разберёшь и не занесёшь решение в translations.json.",
        "",
        "| источник | слово/id | поле | причина |",
        "|---|---|---|---|",
    ]
    for r in rows:
        lines.append(f"| {r['label']} | {r['name']} | {r['field']} | {r['reason']} |")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return path

# ═══════════════════════════════════════════════════════════════
# TERMS / SOUNDS (этап E) — автоперевод убран полностью, значения только
# из docs/en-fix/translations-terms-sounds.json, source: manual. Кэш —
# translations.json["TERMS"]/["SOUNDS"] по id (та же логика загрузки,
# что и load_manual_rules_batches для RULES, но один файл на обе секции).
# ═══════════════════════════════════════════════════════════════
def load_manual_terms_sounds(translations, warn):
    if translations is None:
        return
    f = SCRIPT_DIR / "docs" / "en-fix" / "translations-terms-sounds.json"
    if not f.exists():
        warn.append(f"{f.name} не найден — TERMS/SOUNDS.en не обновлены")
        return
    try:
        data = json.loads(f.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as e:
        warn.append(f"{f.name}: не читается ({e}) — TERMS/SOUNDS.en не обновлены")
        return
    for section in ("TERMS", "SOUNDS"):
        cache = translations[section]
        n = 0
        for tid, entry in data.get(section, {}).items():
            if entry.get("source") and entry.get("source") != "manual":
                warn.append(f"{f.name}: {section}.{tid} без source=manual — пропущена")
                continue
            merged = dict(entry)
            merged["source"] = "manual"
            cache[tid] = merged
            n += 1
        print(f"  ✓ {section} manual: {n} записей из {f.name}")

def apply_manual_terms(terms, translations, warn):
    if translations is None:
        return
    cache = translations["TERMS"]
    without_en = 0
    for t in terms:
        cached = cache.get(t.get("id"))
        if cached and cached.get("source") == "manual":
            if cached.get("en"):
                t["en"] = cached["en"]
            if cached.get("noteEn"):
                t["noteEn"] = cached["noteEn"]
        if not t.get("en"):
            without_en += 1
    if without_en:
        warn.append(f"terms: {without_en} записей без en")

def apply_manual_sounds(sounds, translations, warn):
    if translations is None:
        return
    cache = translations["SOUNDS"]
    for s in sounds:
        cached = cache.get(s.get("id"))
        if cached and cached.get("source") == "manual":
            for en_field in ("comboEn", "pronunciationEn", "exampleEn"):
                if cached.get(en_field):
                    s[en_field] = cached[en_field]

def load_manual_vocab_skip(all_vocab, translations, warn):
    """docs/en-fix/translations-vocab-skip.json — слова, размеченные вручную
    после того как арбитр на Haiku пометил их skip. Файл ключуется по id из
    data.js; кэш translations.json['VOCAB'] — по vocab_hash источника, поэтому
    id → hash сопоставляем здесь по all_vocab.

    source: "manual" (не перезаписывается никакими флагами, GUARDRAILS #26);
    нормализация НЕ применяется — значения в файле уже по конвенциям
    (глаголы 'to X', существительные строчная). Мержим поверх существующей
    записи кэша: en/altEn/noteEn/source из файла перекрывают, прочее
    (напр. _src_note/noteEn от этапа D) сохраняется."""
    if translations is None:
        return
    f = SCRIPT_DIR / "docs" / "en-fix" / "translations-vocab-skip.json"
    if not f.exists():
        warn.append(f"{f.name} не найден — ручные VOCAB.skip не загружены")
        return
    try:
        data = json.loads(f.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as e:
        warn.append(f"{f.name}: не читается ({e}) — ручные VOCAB.skip не загружены")
        return
    by_id = {it["id"]: it for it in all_vocab if it.get("id")}
    cache = translations["VOCAB"]
    today = date.today().isoformat()
    n = 0
    for vid, entry in data.get("VOCAB", {}).items():
        if entry.get("source") and entry.get("source") != "manual":
            warn.append(f"{f.name}: {vid} без source=manual — пропущен")
            continue
        item = by_id.get(vid)
        if item is None:
            warn.append(f"{f.name}: id {vid} нет в словаре (переехал/удалён?) — пропущен")
            continue
        h = vocab_hash(item)
        merged = dict(cache.get(h) or {})
        merged.update(entry)
        merged["source"] = "manual"
        merged.setdefault("de", item.get("de"))
        if entry.get("altEn") and not entry.get("altEnSource"):
            merged["altEnSource"] = "manual"
        merged["date"] = today
        cache[h] = merged
        n += 1
    print(f"  ✓ VOCAB skip manual: {n} записей из {f.name}")

# ─── VOCAB.exampleEn — тот же VOCAB-кэш и тот же хэш, что у слова (exampleDe
# и так часть хэша, см. vocab_hash), но отдельная пара план/прогон: пример
# переводится без контекста и без обратного перевода (короткое предложение,
# не одно слово — сверка EN→DE тут не то же самое, что для слова, и не
# специфицирована в брифе), поэтому не должен зависеть от того, прошло ли
# слово обратный перевод.
def compute_example_plan(all_vocab, translations, warn):
    candidates = []
    cache_hits = 0
    if translations is None:
        return candidates, cache_hits, 0
    vocab_cache = translations["VOCAB"]
    for item in all_vocab:
        if not item.get("exampleDe"):
            continue
        cached = vocab_cache.get(vocab_hash(item))
        if cached and cached.get("exampleEn"):
            cache_hits += 1
            item["exampleEn"] = cached["exampleEn"]
        else:
            candidates.append(item)
    char_count = sum(len(it["exampleDe"]) for it in candidates)
    return candidates, cache_hits, char_count

def run_example_translation(candidates, vocab_cache, warn):
    if not candidates:
        return 0
    if not DEEPL_API_KEY:
        warn.append("DEEPL_API_KEY не задан — VOCAB.exampleEn не переведён")
        return 0
    print(f"  → VOCAB.exampleEn (exampleDe, DE→EN): {len(candidates)} записей...")
    today = date.today().isoformat()
    BATCH = 50
    ok = 0
    for i in range(0, len(candidates), BATCH):
        chunk = candidates[i:i + BATCH]
        try:
            translated = _deepl_translate_batch([it["exampleDe"] for it in chunk], DEEPL_API_KEY, source_lang="DE")
        except Exception as e:
            warn.append(f"DeepL: ошибка перевода VOCAB.exampleDe батч {i}-{i + len(chunk)}: {e}")
            continue
        for item, en_text in zip(chunk, translated):
            item["exampleEn"] = en_text
            entry = vocab_cache.setdefault(vocab_hash(item), {})
            entry["de"] = item["de"]
            entry["exampleEn"] = en_text
            entry.setdefault("source", "deepl")  # не трогать, если запись уже "manual"
            entry.setdefault("date", today)
            ok += 1
    print(f"  ✓ VOCAB.exampleEn: переведено {ok}/{len(candidates)}")
    return ok

# ─── CONJUGATIONS.en / REGEL_VERBS.en — свой раздел в translations.json,
# хэш по verb (единственный источник, без контекста/обратного перевода —
# см. BRIEF §4B: там только про VOCAB.en).
def compute_verb_field_plan(items, field, cache, warn):
    candidates = []
    cache_hits = 0
    for item in items:
        val = item.get(field)
        if not val:
            continue
        h = simple_hash(val)
        cached = cache.get(h)
        if cached and cached.get("en"):
            cache_hits += 1
            item["en"] = cached["en"]
        else:
            candidates.append(item)
    char_count = sum(len(it[field]) for it in candidates)
    return candidates, cache_hits, char_count

def run_verb_field_translation(candidates, field, cache, warn, label):
    if not candidates:
        return 0
    if not DEEPL_API_KEY:
        warn.append(f"DEEPL_API_KEY не задан — {label}.en не переведён")
        return 0
    print(f"  → {label}.en ({field}, DE→EN): {len(candidates)} записей...")
    today = date.today().isoformat()
    BATCH = 50
    ok = 0
    for i in range(0, len(candidates), BATCH):
        chunk = candidates[i:i + BATCH]
        try:
            translated = _deepl_translate_batch([it[field] for it in chunk], DEEPL_API_KEY, source_lang="DE")
        except Exception as e:
            warn.append(f"DeepL: ошибка перевода {label}.{field} батч {i}-{i + len(chunk)}: {e}")
            continue
        for item, en_text in zip(chunk, translated):
            en_text = normalize_verb_en(en_text)
            item["en"] = en_text
            cache[simple_hash(item[field])] = {field: item[field], "en": en_text,
                                                "source": "deepl", "date": today}
            ok += 1
    print(f"  ✓ {label}.en: переведено {ok}/{len(candidates)}")
    return ok

# ─── VOCAB.altEn — этап G. Раньше был единственным полем на старом кэше
# (_parse_old_array_by_id: диф по id против предыдущего data.js). id перевыдаётся
# каждый прогон → кэш промахивался → 289 значений уходили в DeepL всякий раз,
# результат нестабилен. Теперь — тот же механизм, что en/exampleEn (этап B):
# кэш в translations.json["VOCAB"] по vocab_hash, инвалидация по _src_altEn.
# Источник: altDe (DE→EN) либо altRu (RU→EN, помечается altEnSource="deepl-ru").
# Обратного перевода нет ни на одной ветке (решение по §4 брифа): altDe —
# выверенные вручную формы без многозначности, детектор давал только ложные
# срабатывания; altRu защищён контекстом (de как якорь) и пометкой deepl-ru.
def _alt_en_source(item):
    """Ветка источника altEn (BRIEF этап G §2.2). Возвращает
    (branch|None, source_lang|None, source_value|None)."""
    if item.get("altDe"):
        return "altDe", "DE", item["altDe"]
    if item.get("altRu"):
        return "altRu", "RU", item["altRu"]
    return None, None, None

def build_alt_context(item):
    """Контекст для перевода altEn — не тарифицируется. de как якорь: для
    altRu-ветки это главный дизамбигуатор (перевод не должен уходить в
    русский пивот), для altDe — подтверждение синонимии. Плюс артикль,
    часть речи, exampleDe — как для основного слова на этапе B."""
    parts = []
    if item.get("de"):
        parts.append(item["de"])
    article = GENDER_ARTICLE.get(item.get("gender"))
    if article:
        parts.append(article)
    pos_label = POS_DE_LABEL.get(item.get("pos"))
    if pos_label:
        parts.append(pos_label)
    if item.get("exampleDe"):
        parts.append(item["exampleDe"])
    return " | ".join(parts) if parts else None

def compute_alt_en_plan(all_vocab, translations, warn):
    """Без API: делит vocab на кэш-хиты altEn и кандидатов. Кэш —
    translations.json['VOCAB'] по vocab_hash, инвалидация по _src_altEn
    (значение источника на момент перевода — совпало → в DeepL не идём,
    даже если это закэшированный флаг). altEnSource=='manual' не
    пересчитывается никогда (GUARDRAILS #26). Возвращает
    (candidates, cache_hits, flagged_cached, skipped, chars_de, chars_ru)."""
    candidates, cache_hits, flagged_cached, skipped = [], 0, 0, 0
    chars_de = chars_ru = 0
    if translations is None:
        return candidates, cache_hits, flagged_cached, skipped, 0, 0
    vcache = translations["VOCAB"]
    for item in all_vocab:
        branch, _src_lang, src_val = _alt_en_source(item)
        if branch is None:
            skipped += 1
            continue
        entry = vcache.get(vocab_hash(item)) or {}
        if entry.get("altEnSource") == "manual":
            if entry.get("altEn"):
                item["altEn"] = entry["altEn"]
            cache_hits += 1
            continue
        if entry.get("_src_altEn") == src_val:
            if entry.get("altEn"):
                en_val = entry["altEn"]
                renorm = normalize_vocab_en(en_val, item.get("pos"))
                if renorm != en_val:  # нормализация поменялась — локально, без API
                    entry["altEn"] = renorm
                    en_val = renorm
                item["altEn"] = en_val
                cache_hits += 1
            else:  # закэшированный флаг — источник не менялся, заново не гоняем
                flagged_cached += 1
            continue
        candidates.append(item)
        if branch == "altDe":
            chars_de += len(src_val)
        else:
            chars_ru += len(src_val)
    return candidates, cache_hits, flagged_cached, skipped, chars_de, chars_ru

def _cache_alt_en(vocab_cache, item, en_text, src_val, alt_source, today):
    entry = vocab_cache.setdefault(vocab_hash(item), {})
    entry.setdefault("de", item.get("de"))
    entry["altEn"] = en_text
    entry["_src_altEn"] = src_val
    entry["altEnSource"] = alt_source
    entry.setdefault("date", today)
    entry.pop("altEnStatus", None)
    entry.pop("altEnFlag", None)

def _cache_alt_en_flag(vocab_cache, item, src_val, reason, today):
    """Кэшируем исход флага (кириллица) по _src_altEn — иначе прогон 2 гоняет
    его в DeepL заново и стабильности нет по определению. Инвалидируется
    сменой источника или ручным разбором (altEnSource: manual). API-ошибки
    НЕ кэшируем — транзиентны."""
    entry = vocab_cache.setdefault(vocab_hash(item), {})
    entry.setdefault("de", item.get("de"))
    entry["_src_altEn"] = src_val
    entry["altEnStatus"] = "flagged"
    entry["altEnFlag"] = reason
    entry.setdefault("date", today)
    entry.pop("altEn", None)
    entry.pop("altEnSource", None)

def run_alt_en_translation(candidates, vocab_cache, warn):
    """altDe: DE→EN + context. altRu: RU→EN + context. Обратного перевода нет
    ни на одной ветке (решение по §4 брифа): altDe — 42 выверенные вручную
    формы, в основном женский род профессий, многозначности нет и детектор
    давал только ложные срабатывания (12/12); altRu защищён контекстом (de
    как якорь) и пометкой deepl-ru для будущей ревизии.
    Нормализация normalize_vocab_en (глаголы → to X, существительные строчная).
    Кириллица в результате → altEn не пишется, исход флага кэшируется,
    штатный откат на русский. API-ошибка → флаг в отчёт без кэша.
    Возвращает (report_rows, n_de, n_ru)."""
    if not candidates:
        return [], 0, 0
    if not DEEPL_API_KEY:
        warn.append(f"DEEPL_API_KEY не задан — VOCAB.altEn не переведён ({len(candidates)} записей)")
        return [], 0, 0
    today = date.today().isoformat()
    report_rows = []
    n_de = n_ru = 0

    print(f"  → VOCAB.altEn (по слову, DE/RU→EN + контекст): {len(candidates)}...")
    for i, item in enumerate(candidates, 1):
        branch, src_lang, src_val = _alt_en_source(item)
        context = build_alt_context(item)
        if i > 1:
            time.sleep(0.15)  # троттлинг: 308 одиночных запросов подряд ловили 429
        try:
            raw_en = _deepl_translate_one(src_val, DEEPL_API_KEY, src_lang, "EN", context=context)
        except Exception as e:
            warn.append(f"DeepL: ошибка перевода VOCAB.altEn «{item.get('de')}» ({branch}): {e}")
            report_rows.append({"item": item, "branch": branch, "alt_en": "(ошибка API)",
                                 "reason": "ошибка API", "cached": False})
            continue
        en_text = normalize_vocab_en(raw_en, item.get("pos"))
        if has_cyrillic(en_text):
            _cache_alt_en_flag(vocab_cache, item, src_val, "кириллица в результате", today)
            report_rows.append({"item": item, "branch": branch, "alt_en": en_text,
                                 "reason": "кириллица в результате", "cached": True})
            continue
        item["altEn"] = en_text
        _cache_alt_en(vocab_cache, item, en_text, src_val,
                      "deepl" if branch == "altDe" else "deepl-ru", today)
        if branch == "altDe":
            n_de += 1
        else:
            n_ru += 1
        if i % 100 == 0:
            print(f"    ... {i}/{len(candidates)}")

    print(f"  ✓ VOCAB.altEn: altDe {n_de}, altRu {n_ru}, флагов {len(report_rows)}")
    return report_rows, n_de, n_ru

def write_alt_en_report(report_rows, n_de, n_ru, skipped, cache_hits, flagged_cached, vocab_cache):
    REPORTS_DIR.mkdir(exist_ok=True)
    path = REPORTS_DIR / "en_review_altEn.md"
    by_source = {}
    for e in (vocab_cache or {}).values():
        s = e.get("altEnSource")
        if s or e.get("altEnStatus") == "flagged":
            by_source[s or "flagged"] = by_source.get(s or "flagged", 0) + 1
    total_line = ", ".join(f"{k}: {v}" for k, v in sorted(by_source.items())) or "—"
    lines = [
        "# en_review_altEn — статус перевода VOCAB.altEn (этап G)",
        "",
        f"Прогон: {date.today().isoformat()}.",
        f"Переведено в этом прогоне: altDe {n_de}, altRu {n_ru}. "
        f"Кэш-хиты: {cache_hits} (+ {flagged_cached} закэшированных флагов). "
        f"Пропущено (нет ни altDe, ни altRu): {skipped}. Новых флагов: {len(report_rows)}.",
        "",
        f"Всего в translations.json по altEnSource: {total_line}.",
        "",
        "altEn считается: altDe заполнено → DeepL DE→EN + контекст; иначе altRu → "
        "DeepL RU→EN + контекст (пометка altEnSource: deepl-ru). Обратного перевода нет "
        "(altDe — ручные формы, altRu — сравнивать не с чем).",
        "",
    ]
    if report_rows:
        lines += [
            "Поле `altEn` НЕ записано в data.js (сайт покажет русский). Исход закэширован "
            "(`altEnStatus: flagged`) — заново переводиться не будет, пока не сменится источник "
            "или не занесёшь решение вручную (`altEnSource: manual`, `altEn: ...`).",
            "",
            "| id | de | ветка | источник | altEn (DeepL) | причина | закэширован |",
            "|---|---|---|---|---|---|---|",
        ]
        for r in sorted(report_rows, key=lambda x: (x["branch"], x["item"].get("de") or "")):
            it = r["item"]
            src = it.get("altDe") if r["branch"] == "altDe" else it.get("altRu")
            lines.append(f"| {it.get('id','')} | {it.get('de','')} | {r['branch']} | {src or ''} | "
                          f"{r.get('alt_en','')} | {r.get('reason','')} | "
                          f"{'да' if r.get('cached') else 'нет (транзиент)'} |")
    else:
        lines.append("Новых флагов нет — всё прошло нормализацию и проверку на кириллицу.")

    lines += [
        "",
        "## Предсуществующий дефект (§5 брифа, подтверждён — не чинить в рамках G)",
        "",
        "`vocab_hash` для местоимений сводится к `de`+`pos` (нет `gender`, нет `exampleDe`). "
        "Личные местоимения ich/du/wir/ihr/sie заведены в словаре дважды: как личные и внутри "
        "таблицы притяжательных (p0011–p0017). Две разные записи делят одну ячейку кэша "
        "`translations.json['VOCAB']` и переписывают друг другу поля при каждом прогоне.",
        "",
        "Видимое следствие: запись хэша `9e86d0c90b78e6d7` (`du`) осциллирует между "
        "`noteEn: \"deine\"` (притяжательная, note = чистый немецкий, копируется дословно — "
        "результат корректный) и `noteEn: \"informally (duzen)\"` (личная). Осциллирует только "
        "`du`/`dein` — единственная пара, где `note` заполнено у обеих записей; ich/mein, "
        "sie/ihr, wir/unser, ihr/euer молчат случайно (note только у притяжательной). "
        "altEn это не затрагивает — ни у одной из пар нет `altRu`/`altDe`.",
        "",
        "Корень — таблица притяжательных местоимений внутри словаря; правится переносом её "
        "в раздел правил, а не патчем `vocab_hash`.",
    ]
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return path

def _deepl_usage(api_key):
    """GET /v2/usage — израсходовано символов за период. Для baseline до/после
    прогона (BRIEF этап F §6). Запрос не тарифицируется. Возвращает
    (character_count, character_limit) либо (None, None) при ошибке."""
    if not api_key:
        return None, None
    url = _deepl_base_url(api_key).replace("/translate", "/usage")
    req = urllib.request.Request(url, headers={"Authorization": f"DeepL-Auth-Key {api_key}"})
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            d = json.loads(resp.read().decode("utf-8"))
        return d.get("character_count"), d.get("character_limit")
    except (urllib.error.URLError, ValueError, KeyError):
        return None, None

CHANGELOG_RE = re.compile(r"window\.CHANGELOG\s*=\s*(\[.*\n\]);", re.S)

def _load_changelog(changelog_path, warn):
    """(changelog_list | None). Парсит window.CHANGELOG из changelog.js."""
    if not changelog_path.exists():
        return None
    content = changelog_path.read_text(encoding="utf-8")
    m = CHANGELOG_RE.search(content)
    if not m:
        warn.append("changelog.js: не нашёл window.CHANGELOG — перевод пропущен")
        return None
    try:
        return json.loads(m.group(1))
    except json.JSONDecodeError as e:
        warn.append(f"changelog.js: невалидный JSON ({e}) — перевод пропущен, проверь синтаксис вручную")
        return None

def _changelog_should_translate(e, force_changelog):
    """BRIEF этап F §2.2. manualEn: true — руки прочь при любых флагах (§2.2.4).
    Иначе: переводим, если нет titleEn ИЛИ (--force-changelog и это не ручная
    запись)."""
    if e.get("manualEn"):
        return False
    if not e.get("titleEn"):
        return True
    return bool(force_changelog)

def plan_changelog(changelog_path, force_changelog, warn):
    """Для --dry-run: (записи_на_перевод, символов_в_DeepL). Ни одного
    обращения к API. Символы считаются только по строкам с кириллицей —
    полностью латинские/числовые копируются как есть, без запроса."""
    changelog = _load_changelog(changelog_path, warn)
    if changelog is None:
        return [], 0
    entries = [e for e in changelog
               if e.get("title") and _changelog_should_translate(e, force_changelog)]
    chars = 0
    for e in entries:
        for s in [e.get("title", "")] + (e.get("items") or []):
            if has_cyrillic(s):
                chars += len(s)
    return entries, chars

def _changelog_items_invariants(items, parts):
    """BRIEF этап F §2.3 для itemsEn. Возвращает строку-причину или None.
    len совпадает, кириллицы нет, остатков <x/</x нет, каждый латинский токен
    items[i] дословно в parts[i]."""
    if len(parts) != len(items):
        return f"len(itemsEn)={len(parts)} ≠ len(items)={len(items)}"
    for i, (src, tr) in enumerate(zip(items, parts), 1):
        if has_cyrillic(tr):
            return f"элемент {i}: кириллица в переводе"
        if "<x" in tr or "</x" in tr:
            return f"элемент {i}: остаток тега <x/</x"
        missing = [tok for tok in dict.fromkeys(LATIN_WORD_RE.findall(src)) if tok not in tr]
        if missing:
            return f"элемент {i}: пропали токены: {', '.join(missing[:6])}"
    return None

def translate_changelog(changelog_path, warn, force_changelog=False):
    """titleEn/itemsEn для записей changelog.js — RU→EN через DeepL под защитой
    ignore-тегов (BRIEF этап F).

    changelog.js правится руками, не генерируется из xlsx, отдельного "старого
    прогона" нет. Диф-логика проще: переводим запись, если у неё нет titleEn.
    Флаг --force-changelog переводит заново и те, где titleEn уже есть —
    кроме записей с "manualEn": true (перевод написан человеком, §2.2).

    Механика перевода — та же обобщённая обёртка, что на этапах C/D
    (compute_ignoretag_candidates / translate_ignoretag_field): латиница в
    русском тексте по определению немецкая, оборачивается в <x>…</x>,
    tag_handling=xml, ignore_tags="x". Отправка построчная (каждый элемент
    items — отдельная строка в DeepL): если DeepL вернёт элемент с
    добавленными переводами строк, инвариант "число строк" его отбраковывает
    и элемент считается непереведённым (§2.1).

    Инварианты — на каждую запись отдельно (§2.3). Не прошло → titleEn/itemsEn
    не записывается, на сайте показывается русский, строка идёт в
    reports/en_review_changelog.md. Кэша в translations.json у changelog нет —
    защита от повторного перевода это titleEn + флаг manualEn в самом файле.

    Ручная правка: стёр titleEn у записи — она переведётся заново на следующем
    прогоне. Написал перевод вручную — поставь "manualEn": true, тогда никакой
    флаг его не тронет.
    """
    changelog = _load_changelog(changelog_path, warn)
    if changelog is None:
        return

    to_translate = [e for e in changelog
                    if e.get("title") and _changelog_should_translate(e, force_changelog)]
    report_rows = []
    if not to_translate:
        print("  ✓ changelog: новых/непереведённых записей нет")
        _write_changelog_report(changelog, report_rows, force_changelog)
        return
    if not DEEPL_API_KEY:
        warn.append(f"DEEPL_API_KEY не задан — {len(to_translate)} запис(ей) changelog не переведено")
        _write_changelog_report(changelog, report_rows, force_changelog)
        return

    print(f"  → changelog: {len(to_translate)} записей на перевод"
          f"{' (--force-changelog)' if force_changelog else ''}...")
    before_cnt, before_lim = _deepl_usage(DEEPL_API_KEY)
    if before_cnt is not None:
        print(f"    DeepL /usage до: {before_cnt:,}/{before_lim:,} символов")

    # --force-changelog: сносим прежний перевод у переводимых записей, чтобы
    # провал инвариантов давал откат на русский, а не сохранял старый
    # (немецкое-теряющий) вариант. Резерв — changelog.js.bak (BRIEF §2.4).
    for e in to_translate:
        e.pop("titleEn", None)
        e.pop("itemsEn", None)

    kf = lambda e: id(e)  # changelog не в translations.json — ключ = тождество объекта
    ref_date = {id(e): e.get("date", "") for e in to_translate}

    # 1) titleEn — заголовок как одна строка.
    tcache = {}
    tcands = compute_ignoretag_candidates(to_translate, "title", "titleEn", tcache, kf)
    for r in translate_ignoretag_field(tcands, "title", "titleEn", tcache, kf, warn, "changelog"):
        report_rows.append({"id": r["id"], "date": ref_date.get(r["id"], ""),
                            "field": "title", "reason": r["reason"]})

    # 2) itemsEn — каждый элемент items отдельной строкой. Склейка через \n в
    #    одно поле: translate_ignoretag_field шлёт строки в DeepL по одной, а
    #    инвариант "число строк совпадает" ловит расхождение количества.
    icache = {}
    items_cands = []
    for e in to_translate:
        items = e.get("items") or []
        if not items:
            e["itemsEn"] = []
            continue
        e["_itemsJoined"] = "\n".join(items)
        items_cands.append(e)
    ic = compute_ignoretag_candidates(items_cands, "_itemsJoined", "_itemsEnJoined", icache, kf)
    items_fail_refs = {r["id"]: r["reason"]
                       for r in translate_ignoretag_field(ic, "_itemsJoined", "_itemsEnJoined",
                                                          icache, kf, warn, "changelog")}

    # 3) itemsEn обратно в список + changelog-специфичные инварианты (§2.3).
    for e in items_cands:
        joined = e.pop("_itemsEnJoined", None)
        e.pop("_itemsJoined", None)
        items = e["items"]
        if joined is None:
            reason = items_fail_refs.get(id(e), "перевод не прошёл (см. WARN)")
            report_rows.append({"id": id(e), "date": e.get("date", ""), "field": "items", "reason": reason})
            continue
        parts = joined.split("\n")
        bad = _changelog_items_invariants(items, parts)
        if bad:
            report_rows.append({"id": id(e), "date": e.get("date", ""), "field": "items", "reason": bad})
            continue
        e["itemsEn"] = parts

    # titleEn — добор инварианта на остаток тега (translate_ignoretag_field
    # проверяет кириллицу/токены/строки, но не подстроку "<x").
    for e in to_translate:
        te = e.get("titleEn")
        if te and ("<x" in te or "</x" in te):
            report_rows.append({"id": id(e), "date": e.get("date", ""), "field": "title",
                                "reason": "остаток тега <x/</x в titleEn"})
            del e["titleEn"]

    ok_t = sum(1 for e in to_translate if e.get("titleEn"))
    ok_i = sum(1 for e in items_cands if isinstance(e.get("itemsEn"), list) and e.get("itemsEn"))
    print(f"  ✓ changelog: titleEn {ok_t}/{len(to_translate)}, itemsEn {ok_i}/{len(items_cands)}")

    after_cnt, after_lim = _deepl_usage(DEEPL_API_KEY)
    if after_cnt is not None:
        delta = after_cnt - before_cnt if before_cnt is not None else None
        print(f"    DeepL /usage после: {after_cnt:,}/{after_lim:,} символов"
              + (f" (Δ {delta:+,})" if delta is not None else ""))

    # Переписываем файл целиком (нормализует форматирование заодно).
    ordered_entries = []
    for e in changelog:
        ordered = {}
        for k in ("date", "title", "items", "titleEn", "itemsEn", "manualEn"):
            if k in e:
                ordered[k] = e[k]
        ordered_entries.append(ordered)
    new_content = "window.CHANGELOG = " + json.dumps(ordered_entries, ensure_ascii=False, indent=2) + ";\n"
    changelog_path.write_text(new_content, encoding="utf-8")
    _write_changelog_report(changelog, report_rows, force_changelog)

def _write_changelog_report(changelog, report_rows, force_changelog):
    """BRIEF этап F §5: по каждой записи — дата, заголовок, источник, прошли ли
    инварианты, что не прошло."""
    REPORTS_DIR.mkdir(exist_ok=True)
    path = REPORTS_DIR / "en_review_changelog.md"
    fails = {}
    for r in report_rows:
        f = r.get("field", "")
        f = "items" if "item" in f.lower() else f
        fails.setdefault(r["id"], []).append(f"{f}: {r['reason']}")
    lines = [
        "# en_review_changelog — статус EN-перевода changelog.js",
        "",
        f"Прогон: {date.today().isoformat()}. Всего записей: {len(changelog)}. "
        f"--force-changelog: {'да' if force_changelog else 'нет'}.",
        "",
        "Поле, не прошедшее инварианты, в changelog.js не записано — на сайте показывается русский.",
        "",
        "| # | дата | заголовок | источник | titleEn | itemsEn | len(itemsEn)=len(items) | не прошло |",
        "|---|---|---|---|---|---|---|---|",
    ]
    for i, e in enumerate(changelog, 1):
        src = "manual" if e.get("manualEn") else ("deepl" if e.get("titleEn") else "—")
        items = e.get("items") or []
        ien = e.get("itemsEn")
        has_t = "ok" if e.get("titleEn") else "—"
        has_i = "ok" if isinstance(ien, list) and (ien or not items) else "—"
        if not items:
            lens = "n/a"
        elif isinstance(ien, list):
            lens = "ok" if len(ien) == len(items) else f"{len(ien)}≠{len(items)}"
        else:
            lens = "—"
        why = "; ".join(fails.get(id(e), []))
        title_short = (e.get("title") or "")[:38].replace("|", "/")
        lines.append(f"| {i} | {e.get('date','')} | {title_short} | {src} | {has_t} | {has_i} | {lens} | {why} |")

    # Приёмочные токены (BRIEF §4) — ищем в финальном EN-тексте записей.
    def _en_blob(pred):
        out = []
        for e in changelog:
            if pred(e):
                out.append(e.get("titleEn") or "")
                out.extend(e.get("itemsEn") or [])
        return "\n".join(out)
    acc = [
        ("2026-08-23 «Реорганизация тем словаря»",
         lambda e: e.get("date") == "2026-08-23" and "Реорганизац" in (e.get("title") or ""),
         ["stehen", "liegen", "hängen", "stellen", "setzen"]),
        ("запись с Modalverben im Präteritum",
         lambda e: "Modalverben im Präteritum" in (e.get("title") or "")
                   or any("Modalverben im Präteritum" in x for x in (e.get("items") or [])),
         ["Modalverben im Präteritum"]),
        ("2026-06-21 (Adjektivdeklination)",
         lambda e: e.get("date") == "2026-06-21",
         ["Adjektivdeklination"]),
        ("2026-06-02 (Verben mit Dativ)",
         lambda e: e.get("date") == "2026-06-02",
         ["Verben mit Dativ"]),
        ("2026-07-01 (drei Verben)",
         lambda e: e.get("date") == "2026-07-01",
         []),
    ]
    lines += ["", "## Приёмка — контрольные токены (BRIEF §4)", ""]
    for label, pred, tokens in acc:
        blob = _en_blob(pred)
        if not tokens:
            lines.append(f"- **{label}**: EN-текст — см. запись в таблице выше")
            continue
        miss = [t for t in tokens if t not in blob]
        status = "✓ все на месте" if not miss else f"✗ отсутствуют: {', '.join(miss)}"
        lines.append(f"- **{label}**: {status}")

    # Косметические артефакты DeepL (не потеря немецкого, не провал инвариантов —
    # немецкое на месте; это склейка/лишний пробел у границы ignore-тега и
    # неверный смысл у отдельных русских слов). Инварианты этого не ловят.
    artifact_re = re.compile(r"  +|\S\. [a-z]{2}\b|[a-zäöüß]{2}(?:from|and|the|of)\b")
    flagged = []
    for i, e in enumerate(changelog, 1):
        if e.get("manualEn"):
            continue
        for fld in ("titleEn",):
            v = e.get(fld) or ""
            if artifact_re.search(v):
                flagged.append((i, e.get("date", ""), fld, v))
        for j, v in enumerate(e.get("itemsEn") or [], 1):
            if artifact_re.search(v):
                flagged.append((i, e.get("date", ""), f"itemsEn[{j}]", v))
    lines += ["", "## Косметика DeepL (не блокер — немецкое на месте, инварианты пройдены)", ""]
    if flagged:
        for i, d, fld, v in flagged:
            lines.append(f"- #{i} {d} `{fld}`: {v}")
    else:
        lines.append("- не обнаружено")

    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return path

# ═══════════════════════════════════════════════════════════════
# СЛОЙ 3: русские личные формы наст. времени для VOCAB-глаголов
# (нужно trainer.html/PRAESENS_TILES_TMPLS — там v.ru раньше подставлялся
# как есть, инфинитивом: "Составь: «Я любить.»" вместо "«Я люблю.»").
# Генерируем через pymorphy3 из уже существующего поля `ru` в xlsx —
# без новых колонок, без ручной работы автора базы.
#
# pymorphy2 сюда не встал: он тянет pkg_resources через entry points,
# а в актуальных setuptools (>=81) pkg_resources выпилен физически —
# ImportError при первом же MorphAnalyzer(). pymorphy3 — тот же API
# (тот же .parse()/.inflect()), поддерживаемый форк, без этой проблемы.
# ═══════════════════════════════════════════════════════════════
try:
    import pymorphy3
    _MORPH = pymorphy3.MorphAnalyzer()
except ImportError:
    pymorphy3 = None
    _MORPH = None

RU_PRAESENS_PERSONS = [("1per", "sing"), ("2per", "sing"), ("3per", "sing"),
                        ("1per", "plur"), ("2per", "plur"), ("3per", "plur")]

def generate_ru_forms(ru_text):
    """6 личных форм наст. времени (я/ты/он/мы/вы/они) из русского инфинитива.
    Возвращает список из 6 строк либо None, если сгенерировать не удалось —
    вызывающий код в этом случае должен фолбэкнуться на сырой ru_text
    (несовершенный вид без пары, идиома/фраза, слово вне словаря и т.п.).

    Спрягает только первое слово (голову) — хвост фразы вида "ходить по
    магазинам" переносится как есть после проспрягованной формы: "хожу по
    магазинам". Для омографов совершенного/несовершенного вида с одинаковым
    инфинитивом (находить, выглядеть, уходить...) предпочитает несовершенный
    разбор — у совершенного вида в русском нет наст. времени в принципе
    (inflect(pres) для него всегда вернёт None), а у несовершенного есть."""
    if not _MORPH or not ru_text:
        return None
    parts = ru_text.strip().split(None, 1)
    head = parts[0]
    rest = (" " + parts[1]) if len(parts) > 1 else ""
    parses = [p for p in _MORPH.parse(head) if "INFN" in p.tag]
    if not parses:
        return None
    impf = [p for p in parses if "impf" in p.tag]
    chosen = impf[0] if impf else parses[0]
    forms = [chosen.inflect({"pres", per, num}) for per, num in RU_PRAESENS_PERSONS]
    if any(f is None for f in forms):
        return None
    return [f.word + rest for f in forms]

# ═══════════════════════════════════════════════════════════════
# Утилиты для чтения xlsx
# ═══════════════════════════════════════════════════════════════
def clean(v):
    """None / пустую строку → None, иначе str.strip()."""
    if v is None: return None
    s = str(v).strip()
    return s if s else None

def is_true(v):
    """TRUE / true / 1 / x / yes / да → True. Пусто → False."""
    if v is None: return False
    return str(v).strip().lower() in ("true", "1", "x", "yes", "да")

def is_false(v):
    """Явное FALSE / 0 / no / нет → True (для bool с дефолтом True)."""
    if v is None: return False
    return str(v).strip().lower() in ("false", "0", "no", "нет")

def new_status(v):
    """Колонка `new`: TRUE-подобное → "new" (действительно новая запись),
    UPDATE-подобное → "update" (существующую запись отредактировали и
    перевзвели флаг, чтобы она опять попала в отчёт). На сайте оба статуса
    ведут себя одинаково (item.new = True, попадает в буфер «Neu») — разница
    только в консольном отчёте update_data.py. Пусто/не распознано → None."""
    if v is None: return None
    s = str(v).strip().lower()
    if s in ("true", "1", "x", "yes", "да"): return "new"
    if s in ("update", "upd", "u", "обновлено", "обновление"): return "update"
    return None

def to_int(v):
    if v is None or str(v).strip() == "": return None
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return None

def read_sheet(wb, name, warn=None):
    """Читает лист в список dict-ов. Пропускает строку легенды (опц./обяз.)
    и полностью пустые строки. Ключи — заголовки строки 1."""
    if name not in wb.sheetnames:
        print(f"  ⚠ Лист '{name}' не найден")
        return []
    ws = wb[name]
    row1_vals = [c.value for c in ws[1]]
    non_empty_r1 = [v for v in row1_vals if v is not None]
    # Если строка 1 — легенда (обяз./опц.), заголовки в строке 2
    if non_empty_r1 and all(str(v).strip() in ("опц.", "обяз.") for v in non_empty_r1):
        headers = [c.value for c in ws[2]]
        data_start = 3
    else:
        headers = row1_vals
        data_start = 2
    # dict(zip(headers, row)) ниже молча схлопывает повторяющиеся заголовки —
    # вторая колонка с тем же именем затирает первую. Логику не меняем, но
    # предупреждаем (только для непустых имён; пустые/None — отдельный случай).
    named = [str(h).strip() for h in headers if h is not None and str(h).strip()]
    dupes = sorted({h for h in named if named.count(h) > 1})
    if dupes and warn is not None:
        warn.append(f"{name}: повтор заголовков колонок {dupes} — "
                    f"при чтении остаётся только последняя одноимённая колонка")
    rows = []
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        non_empty = [v for v in row if v is not None]
        if non_empty and all(str(v).strip() in ("опц.", "обяз.") for v in non_empty):
            continue
        if not any(v is not None and str(v).strip() for v in row):
            continue
        rows.append(dict(zip(headers, row)))
    return rows

# ═══════════════════════════════════════════════════════════════
# ДЕКЛАРАТИВНАЯ СХЕМА КОЛОНОК
# Формат: (xlsx_col, out_key, kind)
#   kind: "str"  — строка (None если пусто)
#         "bool" — булев флаг (TRUE/1/x/да)
#         "boolT"— булев с дефолтом True (в JS пишем только явный false)
#         "int"  — целое
#         "list" — pipe-separated → массив строк
# Колонки id / domen / group / new обрабатываются отдельно (общая логика).
# Сюда добавляешь новую колонку — она автоматически поедет в data.js.
# ═══════════════════════════════════════════════════════════════
COMMON_META = [
    ("level", "level", "str"),
    ("priority", "priority", "int"),
    ("source", "source", "str"),
    ("example_de", "exampleDe", "str"),
    ("example_ru", "exampleRu", "str"),
    ("quiz_use", "quizUse", "boolT"),   # годится ли ПРИМЕР для теста-сборки
    ("strict_order", "strictOrder", "boolT"),
    ("warning", "warning", "str"),
    ("label", "label", "str"),
    ("note", "note", "str"),
]

SCHEMA = {
    "nouns": [
        ("de", "de", "str"), ("alt-de", "altDe", "str"),
        ("gender", "gender", "str"), ("ru", "ru", "str"), ("alt-ru", "altRu", "str"),
        ("plural", "plural", "str"), ("alt-plural", "altPlural", "str"),
        ("antonym", "antonym", "str"),
    ] + COMMON_META,
    "adjectives": [
        ("de", "de", "str"), ("alt-de", "altDe", "str"),
        ("ru", "ru", "str"), ("alt-ru", "altRu", "str"),
        ("comparative", "comparative", "str"), ("superlative", "superlative", "str"),
        ("antonym", "antonym", "str"), ("derived_from", "derivedFrom", "str"),
    ] + COMMON_META,
    "adverbs": [
        ("de", "de", "str"), ("alt-de", "altDe", "str"),
        ("ru", "ru", "str"), ("alt-ru", "altRu", "str"),
        ("antonym", "antonym", "str"),
    ] + COMMON_META,
    "phrases": [
        ("de", "de", "str"), ("alt-de", "altDe", "str"),
        ("ru", "ru", "str"), ("alt-ru", "altRu", "str"),
        ("context", "context", "str"),
    ] + COMMON_META,
    "pronouns": [
        ("de", "de", "str"), ("alt-de", "altDe", "str"),
        ("ru", "ru", "str"), ("alt-ru", "altRu", "str"),
        ("kind", "kind", "str"),
        ("case", "case", "str"), ("gender", "gender", "str"),
        ("level", "level", "str"), ("priority", "priority", "int"),
        ("source", "source", "str"), ("note", "note", "str"),
    ],
    "numbers": [
        ("digit", "digit", "str"), ("de", "de", "str"),
        ("alt-de", "altDe", "str"),
        ("ru", "ru", "str"), ("alt-ru", "altRu", "str"),
        ("kind", "kind", "str"), ("transcription", "transcription", "str"),
        ("level", "level", "str"), ("priority", "priority", "int"),
        ("source", "source", "str"), ("note", "note", "str"),
    ],
    "terms": [
        ("de", "term", "str"),          # слово термина теперь в колонке de
        ("plural", "plural", "str"), ("ru", "ru", "str"),
        ("level", "level", "str"), ("priority", "priority", "int"),
        ("source", "source", "str"), ("note", "note", "str"),
    ],
    "sounds": [
        ("combo", "combo", "str"), ("pronunciation", "pronunciation", "str"),
        ("example", "example", "str"), ("translation", "translation", "str"),
        ("note", "note", "str"),
    ],
    "rules": [
        ("title", "title", "str"), ("level", "level", "str"),
        ("content_md", "content_md", "str"), ("examples", "examples", "str"),
        ("note", "note", "str"),
    ],
}

def apply_schema(r, sheet):
    """Применяет SCHEMA[sheet] к строке r, возвращает dict (только непустые поля)."""
    out = {}
    for col, key, kind in SCHEMA[sheet]:
        v = r.get(col)
        if kind == "bool":
            if is_true(v): out[key] = True
        elif kind == "boolT":
            # дефолт True: в data.js пишем поле только если ЯВНО false
            if is_false(v): out[key] = False
        elif kind == "int":
            iv = to_int(v)
            if iv is not None: out[key] = iv
        elif kind == "list":
            cv = clean(v)
            if cv: out[key] = [s.strip() for s in cv.split("|") if s.strip()]
        else:  # str
            cv = clean(v)
            if cv is not None: out[key] = cv
    return out

# ═══════════════════════════════════════════════════════════════
# БИБЛИОТЕКА SVG-ИКОНОК (единственный источник — эмитится в data.js
# как ICON_SVGS, потребители: cheatsheet.html (ic()) и trainer.html
# (компонент <Icon name=.../>). Stroke-style, viewBox 24x24.
# ═══════════════════════════════════════════════════════════════
ICON_LIBRARY = {
    "document": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M6 2h8l4 4v14a2 2 0 0 1-2 2H6a2 2 0 0 1-2-2V4a2 2 0 0 1 2-2z'/><path d='M14 2v4h4'/></svg>",
    "chat": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M4 4h16v12H8l-4 4V4z'/></svg>",
    "wave": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M7 11V6a1.5 1.5 0 0 1 3 0v4'/><path d='M10 10V4.5a1.5 1.5 0 0 1 3 0V10'/><path d='M13 10V5.5a1.5 1.5 0 0 1 3 0V11'/><path d='M16 11.5a1.5 1.5 0 0 1 3 0V14a6 6 0 0 1-6 6h-1a6 6 0 0 1-6-6v-1l-2-3a1.3 1.3 0 0 1 2-1.6L7 11'/></svg>",
    "tv": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><rect x='2' y='5' width='20' height='14' rx='2'/><path d='M8 21h8M12 17v4'/></svg>",
    "broom": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M18 3 8 13'/><path d='M8 13c-2 0-4 2-4 4l6-2z'/><path d='M4 21l4-4'/></svg>",
    "sofa": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M4 12v6h16v-6'/><path d='M4 12a2 2 0 0 1 2-2h12a2 2 0 0 1 2 2'/><path d='M4 12V9a1 1 0 0 1 1-1h1a1 1 0 0 1 1 1v2M17 12V9a1 1 0 0 1 1-1h1a1 1 0 0 1 1 1v2'/><path d='M4 18v2M20 18v2'/></svg>",
    "target": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><circle cx='12' cy='12' r='9'/><circle cx='12' cy='12' r='5'/><circle cx='12' cy='12' r='1'/></svg>",
    "palette": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M12 2a10 10 0 1 0 0 20c1.1 0 2-.9 2-2 0-.5-.2-1-.5-1.4-.3-.4-.5-.8-.5-1.3 0-1.1.9-2 2-2h2.3A5.2 5.2 0 0 0 22 10c0-4.4-4.5-8-10-8z'/><circle cx='7' cy='11' r='1'/><circle cx='7.5' cy='7' r='1'/><circle cx='12' cy='5.5' r='1'/><circle cx='16.5' cy='7.5' r='1'/></svg>",
    "carousel": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M12 2v4M12 8l7 4-7 4-7-4z'/><path d='M5 12v6l7 4 7-4v-6'/></svg>",
    "traffic-light": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><rect x='8' y='2' width='8' height='16' rx='3'/><circle cx='12' cy='6' r='1.3'/><circle cx='12' cy='10' r='1.3'/><circle cx='12' cy='14' r='1.3'/><path d='M9 20h6'/></svg>",
    "car": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M3 17V12l2-6h14l2 6v5'/><path d='M3 17h18'/><circle cx='7' cy='17' r='2'/><circle cx='17' cy='17' r='2'/></svg>",
    "paw": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><circle cx='7' cy='8' r='2'/><circle cx='12' cy='6' r='2'/><circle cx='17' cy='8' r='2'/><path d='M12 12c-4 0-6 3-6 5.5S8 21 12 21s6-1 6-3.5S16 12 12 12z'/></svg>",
    "mountain": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M3 20 9 8l4 6 2-3 6 9z'/></svg>",
    "weather-sun": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><circle cx='12' cy='12' r='4'/><path d='M12 2v2M12 20v2M4.9 4.9l1.4 1.4M17.7 17.7l1.4 1.4M2 12h2M20 12h2M4.9 19.1l1.4-1.4M17.7 6.3l1.4-1.4'/></svg>",
    "users": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><circle cx='9' cy='7' r='3'/><path d='M2 21c0-3.9 3.1-7 7-7s7 3.1 7 7'/><circle cx='17' cy='8' r='2.5'/><path d='M22 21c0-3-2-5.5-5-6.3'/></svg>",
    "user": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><circle cx='12' cy='8' r='4'/><path d='M4 21c0-4.4 3.6-8 8-8s8 3.6 8 8'/></svg>",
    "health": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><rect x='3' y='3' width='18' height='18' rx='2'/><path d='M12 8v8M8 12h8'/></svg>",
    "id-card": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><rect x='2' y='5' width='20' height='14' rx='2'/><circle cx='8' cy='11' r='2'/><path d='M5 17c0-1.7 1.3-3 3-3s3 1.3 3 3M14 9h5M14 13h5M14 17h3'/></svg>",
    "briefcase": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><rect x='2' y='7' width='20' height='13' rx='2'/><path d='M8 7V5a2 2 0 0 1 2-2h4a2 2 0 0 1 2 2v2M2 12h20'/></svg>",
    "couple": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M12 21s-7-4.5-9.5-9C.7 8.6 2 5 5.5 5c2 0 3.3 1.2 4.5 2.7C11.2 6.2 12.5 5 14.5 5c3.5 0 4.8 3.6 3 7-2.5 4.5-9.5 9-9.5 9z'/></svg>",
    "mask": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><circle cx='7' cy='12' r='5'/><circle cx='17' cy='12' r='5'/><path d='M5 10.5c1-1 3-1 4 0M15 10.5c1-1 3-1 4 0M5.5 14.5c1 1.5 3 1.5 4 0M14.5 14.5c1 1.5 3 1.5 4 0'/></svg>",
    "brain": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M9 4a3 3 0 0 0-3 3v1a3 3 0 0 0-2 5 3 3 0 0 0 2 5 3 3 0 0 0 5 1V6a3 3 0 0 0-2-2z'/><path d='M15 4a3 3 0 0 1 3 3v1a3 3 0 0 1 2 5 3 3 0 0 1-2 5 3 3 0 0 1-5 1V6a3 3 0 0 1 2-2z'/></svg>",
    "landmark": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M3 21h18M4 21V10M20 21V10M2 10l10-6 10 6M6 10v6M10 10v6M14 10v6M18 10v6'/></svg>",
    "city": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M4 21V9l5-4v16M4 21h16M13 21V4l6 3v14M9 8h1M9 12h1M9 16h1M17 10h1M17 14h1M17 18h1'/></svg>",
    "home": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M3 9.5 12 3l9 6.5'/><path d='M5 8v11a1 1 0 0 0 1 1h4v-6h4v6h4a1 1 0 0 0 1-1V8'/></svg>",
    "door": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><rect x='6' y='2' width='12' height='20' rx='1'/><circle cx='14' cy='12' r='1'/></svg>",
    "shirt": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M8 3 3 6l2 4 2-1v11a1 1 0 0 0 1 1h8a1 1 0 0 0 1-1V9l2 1 2-4-5-3a3 3 0 0 1-6 0z'/></svg>",
    "cup": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M6 8h12l-1.2 11.5a2 2 0 0 1-2 1.5H9.2a2 2 0 0 1-2-1.5L6 8Z'/><path d='M9 3h6l1 5H8z'/></svg>",
    "apple": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M12 7c-3 0-5 2.5-5 6 0 4 2.5 8 5 8s5-4 5-8c0-3.5-2-6-5-6Z'/><path d='M12 7c0-2 1-4 3-4'/></svg>",
    "pan": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><circle cx='11' cy='12' r='7'/><path d='M18 10h4'/><path d='M8 10a3 3 0 0 1 6 0'/></svg>",
    "pencil": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M12 20h9'/><path d='M16.5 3.5a2.1 2.1 0 0 1 3 3L7 19l-4 1 1-4Z'/></svg>",
    "laptop": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><rect x='4' y='4' width='16' height='11' rx='1'/><path d='M2 19h20'/></svg>",
    "cart": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><circle cx='9' cy='20' r='1.3'/><circle cx='18' cy='20' r='1.3'/><path d='M2 3h2l2.6 12.6a2 2 0 0 0 2 1.6h8.8a2 2 0 0 0 2-1.6L21 8H6'/></svg>",
    "smile": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><circle cx='12' cy='12' r='9'/><path d='M8 14s1.5 2 4 2 4-2 4-2'/><path d='M9 9h.01M15 9h.01'/></svg>",
    "warning": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M12 3 2 20h20L12 3Z'/><path d='M12 10v4M12 17h.01'/></svg>",
    "school": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M12 3 2 8l10 5 10-5-10-5Z'/><path d='M6 10.5V16c0 1.5 3 3 6 3s6-1.5 6-3v-5.5M22 8v6'/></svg>",
    "notes": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M4 3h13l3 3v15H4z'/><path d='M17 3v4h3M8 12h8M8 16h8M8 8h4'/></svg>",
    "clock": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><circle cx='12' cy='12' r='9'/><path d='M12 7v5l4 2'/></svg>",
    "calendar": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><rect x='3' y='4' width='18' height='17' rx='2'/><path d='M16 2v4M8 2v4M3 10h18'/></svg>",
    "sunrise": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M12 3v5'/><path d='M5.6 10.6 7 12M18.4 10.6 17 12M2 18h20M4 18a8 8 0 0 1 16 0'/></svg>",
    "hourglass": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M6 2h12M6 22h12M6 2c0 5 4 7 6 8-2 1-6 3-6 8M18 2c0 5-4 7-6 8 2 1 6 3 6 8'/></svg>",
    "stopwatch": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><circle cx='12' cy='13' r='8'/><path d='M12 9v4l3 2'/><path d='M9 2h6M12 2v3'/></svg>",
    "office": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><rect x='4' y='2' width='16' height='20' rx='1'/><path d='M9 7h1M14 7h1M9 11h1M14 11h1M9 15h1M14 15h1M10 22v-4h4v4'/></svg>",
    "crane": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M4 21V9l8-5v17'/><path d='M12 6h8l-3 5M17 11v10'/></svg>",
    "walk": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><circle cx='13' cy='4' r='2'/><path d='M10 21l1-6-3-2 1-5 4-2 3 3h3'/><path d='M9 13l-3 2v6'/></svg>",
    "book": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M4 5a2 2 0 0 1 2-2h13v16H6a2 2 0 0 0-2 2Z'/><path d='M19 17H6a2 2 0 0 0-2 2'/></svg>",
    "ring": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><circle cx='12' cy='15' r='6'/><path d='M9 9l3-6 3 6'/></svg>",
    "flex": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M3 14c2-3 4-4 6-4 3 0 3 2 5 2s3-2 3-2'/><path d='M14 8c1-2 3-3 5-2 2.5 1.3 2 5-1 7-3 2.5-8 3-11 1'/></svg>",
    "ruler": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M3 17 17 3l4 4L7 21z'/><path d='M9 11l2 2M12 8l2 2M6 14l2 2'/></svg>",
    "ruleset": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M3 3v14a4 4 0 0 0 4 4h14'/><path d='M8 3v10M3 8h10'/></svg>",
    "wood": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><rect x='2' y='9' width='20' height='6' rx='2'/><path d='M6 9v6M11 9v6M16 9v6'/></svg>",
    "coin": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><circle cx='12' cy='12' r='9'/><path d='M12 7v10M9.5 9.5c0-1.4 1.1-2.2 2.5-2.2s2.5.8 2.5 2c0 2.5-5 1.7-5 4.2 0 1.2 1.1 2 2.5 2s2.5-.8 2.5-2'/></svg>",
    "star": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M12 2l3 7h7l-5.5 4.5L18.5 21 12 16.5 5.5 21 7.5 13.5 2 9h7Z'/></svg>",
    "handshake": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M2 12h4l3-3 3 3h3'/><path d='M9 12l3 3 6-6 3 3-8 8-6-6'/></svg>",
    "arrow-right": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M5 12h14M13 6l6 6-6 6'/></svg>",
    "pin": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M12 22s7-7.5 7-13a7 7 0 0 0-14 0c0 5.5 7 13 7 13Z'/><circle cx='12' cy='9' r='2.5'/></svg>",
    "thumbs-up": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M7 22V11l4-9c1.5 0 2 1 2 2v6h6a2 2 0 0 1 2 2.3l-1.4 7A2 2 0 0 1 17.6 22H7Z'/><path d='M7 11H3v11h4'/></svg>",
    "repeat": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M17 2l4 4-4 4'/><path d='M3 11V9a4 4 0 0 1 4-4h14'/><path d='M7 22l-4-4 4-4'/><path d='M21 13v2a4 4 0 0 1-4 4H3'/></svg>",
    "graduation": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M2 9 12 4l10 5-10 5z'/><path d='M6 11v5c0 1.5 3 3 6 3s6-1.5 6-3v-5M22 9v6'/></svg>",
    "newspaper": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><rect x='3' y='5' width='13' height='16' rx='1'/><path d='M16 8h5v11a2 2 0 0 1-2 2H5M6 9h6M6 13h6M6 17h6'/></svg>",
    "key": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><circle cx='7' cy='15' r='4'/><path d='M10 12l9-9M16 6l2 2M19 3l2 2'/></svg>",
    "fog": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M4 8h16M2 12h20M4 16h16M6 20h12'/></svg>",
    "ban": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><circle cx='12' cy='12' r='9'/><path d='M6 6l12 12'/></svg>",
    "link": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M9 17H7a5 5 0 0 1 0-10h2M15 7h2a5 5 0 0 1 0 10h-2M8 12h8'/></svg>",
    "medal": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><circle cx='12' cy='15' r='6'/><path d='M9 10 6 3M15 10l3-7M9 3h6'/><path d='M12 12v6'/></svg>",
    "scroll": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M6 4a2 2 0 0 0-2 2v1a2 2 0 0 0 2 2M6 4h13a2 2 0 0 1 2 2v1a2 2 0 0 1-2 2H6M6 4v14M6 18a2 2 0 0 0 2 2h11a2 2 0 0 0 2-2v-1a2 2 0 0 0-2-2H6'/></svg>",
    "letter-a": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M4 20 10 4h4l6 16'/><path d='M7.5 14h9'/></svg>",
    "letter-b": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M6 4h8a4 4 0 0 1 0 8H6zM6 12h9a4 4 0 0 1 0 8H6Z'/></svg>",
    "hash": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M5 9h14M5 15h14M10 3 8 21M16 3l-2 18'/></svg>",
    "box": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M21 8 12 3 3 8l9 5 9-5Z'/><path d='M3 8v9l9 5 9-5V8M12 13v9'/></svg>",
    "bolt": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M13 2 4 14h6l-1 8 9-12h-6z'/></svg>",
    "book-open": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M2 3h6a4 4 0 0 1 4 4v14a3 3 0 0 0-3-3H2z'/><path d='M22 3h-6a4 4 0 0 0-4 4v14a3 3 0 0 1 3-3h7z'/></svg>",
    "waveform": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M2 12h2l2-7 3 14 3-10 2 6 2-3h6'/></svg>",
    "book-check": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M4 5a2 2 0 0 1 2-2h13v16H6a2 2 0 0 0-2 2Z'/><path d='M8 9l2 2 4-4'/></svg>",
    "flame": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M12 2c1 3-3 4-3 8a3 3 0 0 0 6 0c1 1 2 2 2 4a5 5 0 0 1-10 0c0-5 5-6 5-12Z'/></svg>",
    "sparkle": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M12 2l1.5 5.5L19 9l-5.5 1.5L12 16l-1.5-5.5L5 9l5.5-1.5Z'/></svg>",
    "shuffle": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M2 6h4l12 12'/><path d='M14 6h6M18 3l4 3-4 3'/><path d='M2 18h4l3-5'/></svg>",
    "question": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><circle cx='12' cy='12' r='9'/><path d='M9.5 9a2.5 2.5 0 0 1 5 0c0 2-2.5 2-2.5 4.5'/><path d='M12 17h.01'/></svg>",
    "spa": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M12 21c4-2 7-5 7-10a7 7 0 0 0-7-7 7 7 0 0 0-7 7c0 5 3 8 7 10Z'/><path d='M12 21V9'/></svg>",
    "bulb": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M9 18h6M10 21h4'/><path d='M12 3a6 6 0 0 0-4 10.5c.7.6 1 1.5 1 2.5h6c0-1 .3-1.9 1-2.5A6 6 0 0 0 12 3Z'/></svg>",
    "bus": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><rect x='3' y='5' width='18' height='12' rx='2'/><path d='M3 11h18'/><circle cx='7.5' cy='17.5' r='1.5'/><circle cx='16.5' cy='17.5' r='1.5'/></svg>",
    "gamepad": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><rect x='2' y='7' width='20' height='11' rx='5'/><path d='M7 10v4M5 12h4'/><circle cx='16' cy='11' r='1'/><circle cx='18' cy='14' r='1'/></svg>",
    "cloud-rain": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M17 13a4 4 0 0 0-1-7.9A6 6 0 0 0 4.5 9 4 4 0 0 0 5 17h12Z'/><path d='M8 19l-1 2M12 19l-1 2M16 19l-1 2'/></svg>",
    "music": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M9 18V5l11-2v13'/><circle cx='6' cy='18' r='3'/><circle cx='17' cy='16' r='3'/></svg>",
    "wine": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M8 3h8l-1 8a3 3 0 0 1-6 0Z'/><path d='M12 14v7M9 21h6'/></svg>",
    "bed": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M3 18v-6a2 2 0 0 1 2-2h14a2 2 0 0 1 2 2v6'/><path d='M3 18v3M21 18v3M3 12V6h6v6'/></svg>",
    "monitor": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><rect x='3' y='4' width='18' height='12' rx='1'/><path d='M8 20h8M12 16v4'/></svg>",
    "archive": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><rect x='3' y='4' width='18' height='4' rx='1'/><path d='M4 8v10a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8M10 13h4'/></svg>",
    "shower": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M4 12a8 8 0 0 1 14-5'/><path d='M18 3l3 3'/><path d='M4 12h16'/><path d='M8 16v2M12 16v2M16 16v2M6 20v1M10 20v1M14 20v1M18 20v1'/></svg>",
    "toy": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><circle cx='12' cy='8' r='4'/><path d='M9 6l-1-2M15 6l1-2'/><path d='M6 21c0-4 3-7 6-7s6 3 6 7'/></svg>",
    "leaf": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M11 20A7 7 0 0 1 4 13c0-6 5-10 15-11 0 10-4 15-11 15Z'/><path d='M4 20c3-3 5-6 15-15'/></svg>",
    "dining": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M7 2v9M4 2v5a2 2 0 0 0 2 2M10 2v5a2 2 0 0 1-2 2M17 2c-2 0-3 2-3 5s1 4 3 4M17 2v20'/></svg>",
    "bag": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M6 7h12l1 14H5Z'/><path d='M9 7a3 3 0 0 1 6 0'/></svg>",
    "thermometer": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M12 3a2 2 0 0 0-2 2v9a4 4 0 1 0 4 0V5a2 2 0 0 0-2-2Z'/><path d='M12 14v-6'/></svg>",
    "chart": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M4 20V10M10 20V4M16 20v-7M22 20H2'/></svg>",
    "history": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><circle cx='12' cy='13' r='8'/><path d='M12 9v4l3 2'/><path d='M3 8a9 9 0 0 1 2-3M3 8V4M3 8h4'/></svg>",
    "scale": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M12 3v18M5 8h14'/><path d='M5 8l-3 6a3 3 0 0 0 6 0ZM19 8l-3 6a3 3 0 0 0 6 0Z'/><path d='M8 21h8'/></svg>",
    "clipboard": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><rect x='5' y='4' width='14' height='17' rx='2'/><rect x='9' y='2' width='6' height='4' rx='1'/><path d='M9 12h6M9 16h6'/></svg>",
    "check": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M20 6 9 17l-5-5'/></svg>",
    "cross": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M18 6 6 18M6 6l12 12'/></svg>",
    "wrench": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M14.7 6.3a4 4 0 1 0-5.4 5.4L3 18l3 3 6.3-6.3a4 4 0 0 0 5.4-5.4l-2.8 2.8-2-2Z'/></svg>",
    "gear": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><circle cx='12' cy='12' r='3'/><path d='M12 2v3M12 19v3M4.2 4.2l2.1 2.1M17.7 17.7l2.1 2.1M2 12h3M19 12h3M4.2 19.8l2.1-2.1M17.7 6.3l2.1-2.1'/></svg>",
    "rewind": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M11 19 2 12l9-7v14Z'/><path d='M22 19 13 12l9-7v14Z'/></svg>",
    "dice": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><rect x='3' y='3' width='18' height='18' rx='3'/><circle cx='8' cy='8' r='1.3'/><circle cx='16' cy='8' r='1.3'/><circle cx='8' cy='16' r='1.3'/><circle cx='16' cy='16' r='1.3'/><circle cx='12' cy='12' r='1.3'/></svg>",
    "party": "<svg width='1em' height='1em' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M5.8 11.3 2 22l10.7-3.8'/><path d='M11.5 12.5c2-2 5-6 8.5-8.5'/><path d='M16 8c-2 2-6 5-8.5 8.5'/><path d='M4 3h.01M22 8h.01M15 2h.01M22 20h.01M22 14h.01'/></svg>",
}

# ═══════════════════════════════════════════════════════════════
# НЕМЕЦКИЕ ПОДПИСИ ДЛЯ ФРОНТА (BLOCKS / TOPIC_TITLES / TAB_TITLES)
# ─────────────────────────────────────────────────────────────────
# Маппинг ключей domen:group → (icon_key, deutsch). icon_key — ключ
# в ICON_LIBRARY (SVG вместо эмодзи). Хардкод-фоллбэк.
# Если в Settings-XLSX появятся колонки label_de/icon — они
# автоматически переопределят значения отсюда (см. label_for_pair).
# Дописывай сюда новые пары при расширении таксономии.
# ═══════════════════════════════════════════════════════════════
GERMAN_LABELS = {
    # ─── nouns ─────────────────────────────────────────────────
    "communication:docs":       ("document", "Dokumente"),
    "communication:general":    ("chat", "Kommunikation"),
    "communication:greetings":  ("wave", "Begrüßung"),
    "communication:media":      ("tv", "Medien"),
    "home:cleaning":            ("broom", "Putzen"),
    "home:furniture":           ("sofa", "Möbel"),
    "leisure:activities":       ("target", "Aktivitäten"),
    "leisure:hobby":            ("palette", "Hobby"),
    "leisure:places":           ("carousel", "Freizeitorte"),
    "movement:traffic":         ("traffic-light", "Verkehr"),
    "movement:transport":       ("car", "Transport"),
    "nature:animals":           ("paw", "Tiere"),
    "nature:landscape":         ("mountain", "Landschaft"),
    "nature:weather":           ("weather-sun", "Wetter"),
    "people:family":            ("users", "Familie"),
    "people:health":            ("health", "Gesundheit"),
    "people:identity":          ("id-card", "Identität"),
    "people:professions":       ("briefcase", "Berufe"),
    "people:relations":         ("couple", "Beziehungen"),
    "people:roles":             ("mask", "Rollen"),
    "people:traits":            ("brain", "Eigenschaften"),
    "place:buildings":          ("landmark", "Gebäude"),
    "place:city":               ("city", "Stadt"),
    "place:housing":            ("home", "Wohnung"),
    "place:rooms":              ("door", "Zimmer"),
    "products:clothing":        ("shirt", "Kleidung"),
    "products:drinks":          ("cup", "Getränke"),
    "products:food":            ("apple", "Lebensmittel"),
    "products:kitchenware":     ("pan", "Geschirr"),
    "products:stationery":      ("pencil", "Schreibwaren"),
    "products:tech":            ("laptop", "Technik"),
    "shopping:general":         ("cart", "Einkaufen"),
    "state:emotions":           ("smile", "Gefühle"),
    "state:safety":             ("warning", "Sicherheit"),
    "study:places":             ("school", "Lernorte"),
    "study:writing":            ("notes", "Schreiben"),
    "time:clock":               ("clock", "Uhrzeit"),
    "time:dates":               ("calendar", "Datum"),
    "time:dayparts":            ("sunrise", "Tageszeiten"),
    "time:periods":             ("hourglass", "Zeitabschnitte"),
    "work:general":             ("briefcase", "Arbeit"),
    "work:office":              ("office", "Büro"),
    "work:places":              ("crane", "Arbeitsorte"),
    # ─── verbs (доп.) ──────────────────────────────────────────
    "grammar:modal":            ("target", "Modalverben"),
    "home:chores":              ("broom", "Hausarbeit"),
    "movement:general":         ("walk", "Bewegung"),
    "products:cooking":         ("pan", "Kochen"),
    "study:general":            ("book", "Lernen"),
    "time:duration":            ("stopwatch", "Dauer"),
    # ─── adjectives ────────────────────────────────────────────
    "communication:style":      ("chat", "Stil"),
    "movement:speed":           ("bolt", "Tempo"),
    "people:marital":           ("ring", "Familienstand"),
    "people:physical":          ("flex", "Aussehen"),
    "place:size":               ("ruler", "Größe"),
    "products:materials":       ("wood", "Materialien"),
    "shopping:price":           ("coin", "Preis"),
    "state:colors":             ("palette", "Farben"),
    "state:qualities":          ("star", "Qualität"),
    "time:chronology":          ("calendar", "Zeitfolge"),
    # ─── adverbs ───────────────────────────────────────────────
    "communication:agreement":  ("handshake", "Zustimmung"),
    "movement:direction":       ("arrow-right", "Richtung"),
    "place:location":           ("pin", "Ort"),
    "state:certainty":          ("target", "Sicherheit"),
    "state:feelings":           ("thumbs-up", "Befinden"),
    "time:frequency":           ("repeat", "Häufigkeit"),
    # ─── phrases (доп.) ────────────────────────────────────────
    "communication:questions":  ("question", "Fragen"),
    "communication:wellbeing":  ("spa", "Befinden"),
    "study:classroom":          ("graduation", "Im Unterricht"),
    # ─── rules (домен grammar) ─────────────────────────────────
    "grammar:pronouns":         ("user", "Pronomen"),
    "grammar:wordorder":        ("ruleset", "Satzbau"),
    "grammar:verbs":            ("bolt", "Verben"),
    "grammar:cases":            ("target", "Kasus"),
    "grammar:articles":         ("newspaper", "Artikel"),
    "grammar:wfragen":          ("question", "W-Fragen"),
    "grammar:adjectives":       ("palette", "Adjektive"),
    # ─── pronouns (домен grammar / специальные) ────────────────
    "grammar:personal":         ("user", "Personalpronomen"),
    "grammar:possessive":       ("key", "Possessivpronomen"),
    "grammar:indefinite":       ("fog", "Indefinitpronomen"),
    "grammar:negation":         ("ban", "Negation"),
    "grammar:particles":        ("target", "Partikeln"),
    "grammar:prepositions":     ("link", "Präpositionen"),
    # ─── numbers (доп.) ────────────────────────────────────────
    "quantity:ordinal":         ("medal", "Ordinalzahlen"),
    # ─── phonetics (звуки) ─────────────────────────────────────
    "phonetics:rules":          ("scroll", "Regeln"),
    "phonetics:vowels":         ("letter-a", "Vokale"),
    "phonetics:consonants":     ("letter-b", "Konsonanten"),
    # ─── специальные (не из Settings) ──────────────────────────
    "quantity:cardinal":        ("hash", "Kardinalzahlen"),
    "pron:personal":            ("user", "Personalpronomen"),
    "pron:negation":            ("ban", "Negation"),
    "new:new-nouns":            ("box", "Neue Substantive"),
    "new:new-verbs":             ("bolt", "Neue Verben"),
    "new:new-adj":               ("palette", "Neue Adjektive"),
    "new:new-adv":               ("stopwatch", "Neue Adverbien"),
    "new:new-phrases":           ("chat", "Neue Phrasen"),
    "new:new-pron":              ("user", "Neue Pronomen"),
    "new:new-nums":              ("hash", "Neue Zahlen"),
    "new:new-terms":             ("book-open", "Neue Begriffe"),
    "new:new-sounds":            ("waveform", "Neue Laute"),
    "new:new-rules":             ("ruleset", "Neue Regeln"),
    "perfekt:regel":            ("book-check", "Regelmäßig"),
    "perfekt:unregel":          ("flame", "Unregelmäßig"),
    # ─── ★ новые домены (пар пока нет в базе — заранее, под будущие данные) ──
    "basic:general":            ("gear",        "Grundverben"),
    "basic:mental":             ("bulb",        "Denken & Wissen"),
    "basic:senses":             ("waveform",    "Wahrnehmung"),
    "basic:actions":            ("box",         "Handlungen"),
    "basic:events":             ("flame",       "Ereignisse"),
    "money:general":            ("coin",        "Geld & Preise"),
    "money:banking":            ("landmark",    "Bank & Konto"),
    "money:income":             ("chart",       "Einkommen"),
    "admin:documents":          ("document",    "Dokumente"),
    "admin:contracts":          ("scroll",      "Verträge"),
    "admin:insurance":          ("clipboard",   "Versicherungen"),
    "admin:general":            ("id-card",     "Behördengänge"),
    "society:general":          ("users",       "Gesellschaft"),
    "quality:size":             ("ruler",       "Größe"),
    "quality:condition":        ("check",       "Zustand"),
    "quality:sensory":          ("thermometer", "Sinneseindrücke"),
    "quality:evaluation":       ("star",        "Bewertung"),
    "quality:difference":       ("scale",       "Vergleich"),
    "quality:colors":           ("palette",     "Farben"),
    "quality:patterns":         ("shuffle",     "Muster"),
    # ★ новые группы в старых доменах
    "communication:speech":     ("chat",        "Sprechen"),
    "communication:requests":   ("handshake",   "Bitten & Raten"),
    "communication:info":       ("newspaper",   "Auskunft"),
    "home:living":              ("key",         "Wohnen & Umzug"),
    "home:repair":              ("wrench",      "Reparatur"),
    "home:routine":             ("sunrise",     "Tagesablauf"),
    "place:position":           ("pin",         "Position"),
    "movement:objects":         ("carousel",    "Bewegung von Dingen"),
    "movement:transfer":        ("bag",         "Bringen & Holen"),
    "leisure:games":            ("gamepad",     "Spiele & Sport"),
    "nature:garden":            ("leaf",        "Garten"),
    "work:time":                ("stopwatch",   "Arbeitszeit"),
    "shopping:goods":           ("bag",         "Ware & Bestellung"),
    "study:exams":              ("book-check",  "Prüfungen"),
    "study:events":             ("party",       "Schulveranstaltungen"),
    "state:problems":           ("warning",     "Probleme & Lösungen"),
    "state:necessity":          ("target",      "Notwendigkeit"),
    "quantity:measure":         ("ruler",       "Maße"),
    "quantity:general":         ("hash",        "Menge"),
    # ☆ старый пробел — эти пары были в базе, но рендерились сырым ключом
    "state:abstract":           ("bulb",        "Abstraktes"),
    "home:general":             ("home",        "Haushalt"),
    "home:appliances":          ("monitor",     "Haushaltsgeräte"),
    "home:bathroom":            ("shower",      "Badezimmer"),
    "home:textiles":            ("bed",         "Textilien"),
    "leisure:sport":            ("medal",       "Sport"),
    "place:shops":              ("cart",        "Geschäfte"),
    "nature:plants":            ("leaf",        "Pflanzen"),
    "people:body":              ("flex",        "Körper"),
    "products:general":         ("box",         "Umgang mit Dingen"),
    "shopping:units":           ("scale",       "Mengenangaben"),
    "style:slang":              ("mask",        "Umgangssprache"),
    "communication:text":       ("notes",       "Text & Brief"),
    "grammar:adverbs":          ("shuffle",     "Adverbien"),
    "grammar:wordformation":    ("ruleset",     "Wortbildung"),
}

# РУССКИЕ ПЕРЕВОДЫ (для тултипов на фронте)
# ─────────────────────────────────────────────────────────────────
RUSSIAN_LABELS = {
    # ─── существительные ────────────────────────────────────────
    "communication:docs":       "Документы",
    "communication:general":    "Общение",
    "communication:greetings":  "Приветствия",
    "communication:media":      "СМИ / медиа",
    "home:cleaning":            "Уборка",
    "home:furniture":           "Мебель",
    "leisure:activities":       "Занятия",
    "leisure:hobby":            "Хобби",
    "leisure:places":           "Места отдыха",
    "movement:traffic":         "Дорожное движение",
    "movement:transport":       "Транспорт",
    "nature:animals":           "Животные",
    "nature:landscape":         "Природа / пейзаж",
    "nature:weather":           "Погода",
    "people:family":            "Семья",
    "people:health":            "Здоровье",
    "people:identity":          "Личность / удостоверение",
    "people:professions":       "Профессии",
    "people:relations":         "Отношения",
    "people:roles":             "Роли",
    "people:traits":            "Черты характера",
    "place:buildings":          "Здания",
    "place:city":               "Город",
    "place:housing":            "Жильё",
    "place:rooms":              "Комнаты",
    "products:clothing":        "Одежда",
    "products:drinks":          "Напитки",
    "products:food":            "Еда / продукты",
    "products:kitchenware":     "Посуда",
    "products:stationery":      "Канцелярия",
    "products:tech":            "Техника",
    "shopping:general":         "Покупки",
    "state:emotions":           "Чувства / эмоции",
    "state:safety":             "Безопасность",
    "study:places":             "Места учёбы",
    "study:writing":            "Письмо / текст",
    "time:clock":               "Время (часы)",
    "time:dates":               "Даты / дни недели",
    "time:dayparts":            "Части дня",
    "time:periods":             "Периоды времени",
    "work:general":             "Работа",
    "work:office":              "Офис",
    "work:places":              "Рабочие места",
    # ─── глаголы (доп.) ─────────────────────────────────────────
    "grammar:modal":            "Модальные глаголы",
    "home:chores":              "Домашние дела",
    "movement:general":         "Движение",
    "products:cooking":         "Готовка",
    "study:general":            "Учёба",
    "time:duration":            "Продолжительность",
    "communication:style":      "Стиль общения",
    # ─── прилагательные ─────────────────────────────────────────
    "movement:speed":           "Скорость",
    "people:marital":           "Семейное положение",
    "people:physical":          "Внешность",
    "place:size":               "Размер",
    "products:materials":       "Материалы",
    "shopping:price":           "Цена",
    "state:colors":             "Цвета",
    "state:qualities":          "Качества",
    "time:chronology":          "Хронология",
    # ─── наречия ────────────────────────────────────────────────
    "communication:agreement":  "Согласие / несогласие",
    "movement:direction":       "Направление",
    "place:location":           "Местоположение",
    "state:certainty":          "Уверенность",
    "state:feelings":           "Самочувствие",
    "time:frequency":           "Частота",
    # ─── фразы ──────────────────────────────────────────────────
    "communication:questions":  "Вопросы",
    "communication:wellbeing":  "Самочувствие",
    "study:classroom":          "В классе",
    # ─── грамматика / правила ────────────────────────────────────
    "grammar:pronouns":         "Местоимения",
    "grammar:wordorder":        "Порядок слов",
    "grammar:verbs":            "Глаголы",
    "grammar:cases":            "Падежи",
    "grammar:articles":         "Артикли",
    "grammar:wfragen":          "W-вопросы",
    "grammar:adjectives":       "Прилагательные",
    "grammar:personal":         "Личные местоимения",
    "grammar:possessive":       "Притяжательные местоимения",
    "grammar:indefinite":       "Неопределённые местоимения",
    "grammar:negation":         "Отрицание",
    "grammar:particles":        "Частицы",
    "grammar:prepositions":     "Предлоги",
    # ─── числа / фонетика ───────────────────────────────────────
    "quantity:cardinal":        "Количественные числительные",
    "quantity:ordinal":         "Порядковые числительные",
    "phonetics:rules":          "Правила произношения",
    "phonetics:vowels":         "Гласные",
    "phonetics:consonants":     "Согласные",
    # ─── специальные ────────────────────────────────────────────
    "pron:personal":            "Личные местоимения",
    "pron:negation":            "Отрицание",
    "new:new-nouns":            "Новые существительные",
    "new:new-verbs":            "Новые глаголы",
    "new:new-adj":              "Новые прилагательные",
    "new:new-adv":              "Новые наречия",
    "new:new-phrases":          "Новые фразы",
    "new:new-pron":             "Новые местоимения",
    "new:new-nums":             "Новые числа",
    "new:new-terms":            "Новые термины",
    "new:new-sounds":           "Новые звуки",
    "new:new-rules":            "Новые правила",
    "perfekt:regel":            "Регулярные глаголы",
    "perfekt:unregel":          "Нерегулярные глаголы",
    # ★ новые пары (см. GERMAN_LABELS выше) + ☆ старый пробел
    "basic:general":            "Базовые глаголы",
    "basic:mental":             "Мышление и знание",
    "basic:senses":             "Восприятие",
    "basic:actions":            "Действия с предметами",
    "basic:events":             "События и процессы",
    "money:general":            "Деньги и цены",
    "money:banking":            "Банк и счёт",
    "money:income":             "Доход",
    "admin:documents":          "Документы",
    "admin:contracts":          "Договоры",
    "admin:insurance":          "Страхование",
    "admin:general":            "Дела в ведомствах",
    "society:general":          "Общество",
    "quality:size":             "Размер",
    "quality:condition":        "Состояние предмета",
    "quality:sensory":          "Ощущения",
    "quality:evaluation":       "Оценка",
    "quality:difference":       "Сходство и различие",
    "quality:colors":           "Цвета",
    "quality:patterns":         "Узоры",
    "communication:speech":     "Речь",
    "communication:requests":   "Просьбы и советы",
    "communication:info":       "Справка и информация",
    "home:living":              "Жильё и переезд",
    "home:repair":              "Ремонт",
    "home:routine":             "Режим дня",
    "place:position":           "Положение в пространстве",
    "movement:objects":         "Движение предметов",
    "movement:transfer":        "Принести / забрать",
    "leisure:games":            "Игры и спорт",
    "nature:garden":            "Сад и растения",
    "work:time":                "Рабочее время",
    "shopping:goods":           "Товар и заказ",
    "study:exams":              "Экзамены",
    "study:events":             "Учебные мероприятия",
    "state:problems":           "Проблемы и решения",
    "state:necessity":          "Необходимость",
    "quantity:measure":         "Измерения",
    "quantity:general":         "Количество",
    "state:abstract":           "Абстрактные понятия",
    "home:general":             "Дом / хозяйство",
    "home:appliances":          "Бытовая техника",
    "home:bathroom":            "Ванная",
    "home:textiles":            "Текстиль",
    "leisure:sport":            "Спорт",
    "place:shops":              "Магазины",
    "nature:plants":            "Растения",
    "people:body":              "Тело",
    "products:general":         "Обращение с вещами",
    "shopping:units":           "Меры и упаковка",
    "style:slang":              "Сленг / разговорное",
    "communication:text":       "Текст и письмо",
    "grammar:adverbs":          "Наречия",
    "grammar:wordformation":    "Словообразование",
}

BLOCK_META_RU = {
    "neu":     "Новые слова (буфер)",
    "nums":    "Числа",
    "sounds":  "Произношение",
    "verbs":   "Глаголы",
    "unregel": "Нерегулярные глаголы в Präsens",
    "perfekt": "Прошедшее время (Perfekt)",
    "nouns":   "Существительные",
    "adj":     "Прилагательные",
    "adv":     "Наречия",
    "mestoim": "Местоимения",
    "termin":  "Грамматические термины",
    "rules":   "Правила грамматики",
    "phrases": "Разговорные фразы",
}

# Подписи и цвета верхнеуровневых блоков (POS / спец.).
BLOCK_META = {
    "neu":     {"label": "Neu",            "icon": "sparkle", "color": "var(--c-block-neu)", "desc": "Frische Wörter (Puffer)"},
    "verbs":   {"label": "Verben",         "icon": "bolt",    "color": "var(--c-block-verben)"},
    "nouns":   {"label": "Substantive",    "icon": "box",     "color": "var(--c-block-substantive)"},
    "adj":     {"label": "Adjektive",      "icon": "palette", "color": "var(--c-block-adj)"},
    "adv":     {"label": "Adverbien",      "icon": "shuffle", "color": "var(--c-block-adv)"},
    "mestoim": {"label": "Pronomen",       "icon": "user",    "color": "var(--c-block-pronomen)"},
    "phrases": {"label": "Redewendungen",  "icon": "chat",    "color": "var(--c-block-redewendungen)"},
    "unregel": {"label": "Unregelmäßig",   "icon": "flame",   "color": "var(--c-block-unregel)", "desc": "Unregelmäßige Verben in Präsens", "kind": "conjugations"},
    "nums":    {"label": "Zahlen",         "icon": "hash",    "color": "#16a085"},
    "sounds":  {"label": "Aussprache",     "icon": "waveform", "color": "#9b59b6", "desc": "Ausspracheregeln", "kind": "sounds"},
}

def label_for_pair(domen, group, taxonomy_entry=None):
    """Возвращает {"icon": icon_key, "label": Подпись} для domen:group.
    Override через колонки label_de/label/icon в Settings-XLSX (icon —
    ключ в ICON_LIBRARY, не эмодзи)."""
    if taxonomy_entry:
        de_label = taxonomy_entry.get("label_de") or taxonomy_entry.get("label")
        icon = taxonomy_entry.get("icon")
        if de_label:
            return {"icon": icon or "", "label": de_label}
    key = f"{domen}:{group}"
    icon, name = GERMAN_LABELS.get(key, ("", key))
    return {"icon": icon, "label": name}

def build_block_from_pairs(bid, taxonomy_pairs, used_topics=None):
    """Блок POS: каждая пара domen:group → подблок.
    Если передан used_topics — отфильтровываем подблоки без слов."""
    meta = BLOCK_META[bid]
    block = {"id": bid, "label": meta["label"], "icon": meta.get("icon", ""), "color": meta["color"]}
    if "desc" in meta: block["desc"] = meta["desc"]
    seen = set()
    subs = []
    for p in (taxonomy_pairs or []):
        d, g = p.get("domen"), p.get("group")
        if not (d and g) or (d, g) in seen: continue
        seen.add((d, g))
        topic = f"{d}:{g}"
        if used_topics is not None and topic not in used_topics:
            continue  # пустой подблок — не показываем
        lp = label_for_pair(d, g, p)
        subs.append({
            "id": f"{d}-{g}".replace("/", "-"),
            "label": lp["label"],
            "icon": lp["icon"],
            "topics": [topic],
        })
    if subs: block["subblocks"] = subs
    return block

def build_special_block(bid, subblocks=None):
    """Спец. блок (kind=sounds/terms/rules/conjugations/perfekt или с фикс. подблоками)."""
    meta = BLOCK_META[bid]
    block = {"id": bid, "label": meta["label"], "icon": meta.get("icon", ""), "color": meta["color"]}
    if "desc" in meta: block["desc"] = meta["desc"]
    if "kind" in meta: block["kind"] = meta["kind"]
    if subblocks: block["subblocks"] = subblocks
    return block

def collect_used_topics_by_pos(vocab):
    """Для каждого pos — множество topics, реально встречающихся у слов."""
    out = {}
    for v in vocab:
        pos = v.get("pos")
        for t in v.get("topics", []) or []:
            out.setdefault(pos, set()).add(t)
    return out

def build_blocks(taxonomy, vocab=None):
    """BLOCKS строится из Settings-XLSX + спец. блоков. Порядок важен (UI).
    Если передан vocab — отфильтровываем пустые подблоки.

    Архитектура:
    - В тренажёре 7 блоков: Neu, Verben (+Unregel внутри), Substantive,
      Adjektive, Adverbien, Pronomen, Redewendungen.
    - Sounds остаётся в массиве (kind=sounds), но тренажёр его фильтрует —
      используется только в справочнике.
    - Unregel — сабблок Verben (kind=conjugations).
      trainer.html распознаёт sub.kind и роутит его в свой поток.
    - Блок Perfekt удалён: его функционал (тренировка Partizip II) полностью
      дублируется комбо-кнопкой «Partizip II» на экране select-test-type.
      Контекст задаёт фильтр автоматически: в Verben — все глаголы (regel+unregel),
      в Unregel-сабблоке — только unregel.
    """
    used = collect_used_topics_by_pos(vocab or [])

    # Verben + спец. сабблок Unregel в начале
    verbs_block = build_block_from_pairs("verbs", taxonomy.get("verbs", []), used.get("verb"))
    unregel_sub = {
        "id": "unregel",
        "label": BLOCK_META["unregel"]["label"],
        "icon": BLOCK_META["unregel"].get("icon", ""),
        "color": BLOCK_META["unregel"]["color"],
        "desc": BLOCK_META["unregel"].get("desc", ""),
        "kind": "conjugations",
    }
    verbs_block["subblocks"] = [unregel_sub] + verbs_block.get("subblocks", [])

    return [
        # 1. Neu (буфер новых)
        build_special_block("neu", subblocks=[
            {"id": "nouns", "label": "Substantive", "icon": "box",       "topics": ["new:new-nouns"]},
            {"id": "verbs", "label": "Verben",      "icon": "bolt",      "topics": ["new:new-verbs"]},
            {"id": "adj",   "label": "Adjektive",   "icon": "palette",   "topics": ["new:new-adj"]},
            {"id": "adv",   "label": "Adverbien",   "icon": "stopwatch", "topics": ["new:new-adv"]},
        ]),
        # 2. Aussprache (kind=sounds, только для справочника; trainer фильтрует)
        build_special_block("sounds"),
        # 3. Verben (+ Unregel-сабблок в начале)
        verbs_block,
        # 4. Substantive
        build_block_from_pairs("nouns", taxonomy.get("nouns", []), used.get("noun")),
        # 5. Adjektive
        build_block_from_pairs("adj", taxonomy.get("adjectives", []), used.get("adj")),
        # 6. Adverbien
        build_block_from_pairs("adv", taxonomy.get("adverbs", []), used.get("adv")),
        # 7. Pronomen
        build_block_from_pairs("mestoim", taxonomy.get("pronouns", []), used.get("pron")),
        # 8. Redewendungen
        build_block_from_pairs("phrases", taxonomy.get("phrases", []), used.get("phrase")),
    ]

def build_topic_titles(taxonomy):
    """TOPIC_TITLES — заголовки секций в справочнике. Каждое значение — {icon, label}."""
    out = {}
    for page, pairs in taxonomy.items():
        for p in pairs:
            d, g = p.get("domen"), p.get("group")
            if not (d and g): continue
            out[f"{d}:{g}"] = label_for_pair(d, g, p)
    for k in ("quantity:cardinal", "pron:personal", "pron:negation",
              "new:new-nouns", "new:new-verbs", "new:new-adj", "new:new-adv",
              "new:new-phrases", "new:new-pron", "new:new-nums",
              "new:new-terms", "new:new-sounds", "new:new-rules",
              "perfekt:regel", "perfekt:unregel"):
        icon, name = GERMAN_LABELS.get(k, ("", k))
        out[k] = {"icon": icon, "label": name}
    return out

def build_tab_titles():
    return {bid: {"icon": meta.get("icon", ""), "label": meta["label"]} for bid, meta in BLOCK_META.items()}

def build_ru_titles(taxonomy):
    """TOPIC_TITLES_RU и TAB_TITLES_RU — русские переводы для тултипов."""
    topic_ru = {}
    for pairs in taxonomy.values():
        for p in pairs:
            d, g = p.get("domen"), p.get("group")
            if not (d and g): continue
            key = f"{d}:{g}"
            if key in RUSSIAN_LABELS:
                topic_ru[key] = RUSSIAN_LABELS[key]
    # специальные ключи
    for k in ("quantity:cardinal", "pron:personal", "pron:negation",
              "new:new-nouns", "new:new-verbs", "new:new-adj", "new:new-adv",
              "new:new-phrases", "new:new-pron", "new:new-nums",
              "new:new-terms", "new:new-sounds", "new:new-rules",
              "perfekt:regel", "perfekt:unregel"):
        if k in RUSSIAN_LABELS:
            topic_ru[k] = RUSSIAN_LABELS[k]
    return topic_ru, dict(BLOCK_META_RU)

# ───────────────────────────────────────────────────────────────
# Хелперы domen/group → topics
# ───────────────────────────────────────────────────────────────
def dg(r):
    """Возвращает (domen, group) очищенные или (None, None)."""
    return clean(r.get("domen")), clean(r.get("group"))

def topic_key(domen, group):
    """Канонический ключ темы. domen:group, либо domen, либо None."""
    if domen and group: return f"{domen}:{group}"
    if domen: return domen
    return None

# ═══════════════════════════════════════════════════════════════
# ОБРАБОТКА ЛИСТОВ
# ═══════════════════════════════════════════════════════════════
NEW_TOPIC = {  # буфер «Новые» по pos
    "noun":   "new:new-nouns",
    "verb":   "new:new-verbs",
    "adj":    "new:new-adj",
    "adv":    "new:new-adv",
    "phrase": "new:new-phrases",
    "pron":   "new:new-pron",
    "num":    "new:new-nums",
    "term":   "new:new-terms",
    "sound":  "new:new-sounds",
    "rule":   "new:new-rules",
}

def attach_common(item, r, pos):
    """Добавляет id/domen/group/topics/new к vocab-элементу."""
    rid = clean(r.get("id"))
    if rid: item["id"] = rid
    domen, group = dg(r)
    if domen: item["domen"] = domen
    if group: item["group"] = group
    topics = []
    tk = topic_key(domen, group)
    if tk: topics.append(tk)
    status = new_status(r.get("new"))
    if status:
        item["new"] = True
        if status == "update": item["_newUpdate"] = True
        nt = NEW_TOPIC.get(pos)
        if nt and nt not in topics:
            topics.append(nt)
    item["topics"] = topics

def process_simple_vocab(rows, sheet, pos, warn):
    """nouns / adjectives / adverbs / phrases — общая логика."""
    items = []
    seen = set()
    for r in rows:
        de = clean(r.get("de"))
        if not de: continue
        # legacy: инлайн "X / Y" в de → de + altDe (на случай старых записей)
        alt_inline = None
        if " / " in de and not clean(r.get("alt-de")):
            parts = [p.strip() for p in de.split(" / ")]
            de = parts[0]
            alt_inline = " / ".join(parts[1:]) if len(parts) > 1 else None
        domen, group = dg(r)
        key = (de, domen or "", group or "")
        if key in seen:
            warn.append(f"{sheet}: дубль «{de}» ({domen}/{group})")
            continue
        seen.add(key)
        item = {"de": de}
        item.update(apply_schema(r, sheet))
        if alt_inline and "altDe" not in item:
            item["altDe"] = alt_inline
        item["pos"] = pos
        attach_common(item, r, pos)
        items.append(item)
    return items

def process_verbs(rows, warn):
    """→ (vocab_items, regel_verbs, conjugations)."""
    vocab_items, regel_verbs, conjugations = [], [], []
    seen = set()
    missing_p2, missing_aux = [], []

    for r in rows:
        verb = clean(r.get("de"))
        if not verb: continue
        if verb in seen:
            warn.append(f"verbs: дубль «{verb}»")
            continue
        seen.add(verb)

        vtype = (clean(r.get("type")) or "regel").lower()
        if vtype not in ("regel", "unregel", "composite"):
            vtype = "regel"

        # ── VOCAB-запись глагола (все поля схемы verbs) ──
        vi = {"de": verb}
        vi.update(apply_schema_verb(r))
        vi["pos"] = "verb"
        attach_common(vi, r, "verb")
        vocab_items.append(vi)

        if vtype == "composite":
            continue  # только в VOCAB

        # формы Präsens
        forms = [clean(r.get(c)) for c in ("ich", "du", "er_sie_es", "wir", "ihr", "sie_Sie")]
        forms = [f or "" for f in forms]
        has_all_forms = all(forms)

        aux = (clean(r.get("aux")) or "").lower()
        if aux not in ("haben", "sein", ""):
            aux = ""
        p2 = clean(r.get("partizip2")) or ""
        praet = clean(r.get("praeteritum"))
        modal = is_true(r.get("modal"))

        if vtype == "unregel":
            if not has_all_forms:
                warn.append(f"verbs: unregel без полных форм Präsens — «{verb}» (автоген)")
                forms = conj_all_forms(verb)
            if not p2: missing_p2.append(verb)
            if not aux: missing_aux.append(verb)
            conj = {
                "verb": verb,
                "ru": clean(r.get("ru")) or "",
                "tense": "Präsens",
                "pronouns": ["ich", "du", "er/sie/es", "wir", "ihr", "sie/Sie"],
                "forms": forms,
            }
            if modal: conj["modal"] = True
            if p2: conj["partizip2"] = p2
            if aux: conj["aux"] = aux
            if praet: conj["praeteritum"] = praet
            for fl in ("separable", "reflexive", "impersonal"):
                if is_true(r.get(fl)): conj[fl] = True
            if clean(r.get("case")): conj["case"] = clean(r.get("case"))
            if clean(r.get("level")): conj["level"] = clean(r.get("level"))
            conjugations.append(conj)
        else:  # regel
            if not has_all_forms:
                forms = conj_all_forms(verb)
            if not p2:
                p2 = partizip2_regelmaessig(verb)
            if not aux:
                aux = "haben"
            reg = {"verb": verb, "ru": clean(r.get("ru")) or "",
                   "forms": forms, "partizip2": p2, "aux": aux}
            if praet: reg["praeteritum"] = praet
            for fl in ("separable", "reflexive", "impersonal"):
                if is_true(r.get(fl)): reg[fl] = True
            if clean(r.get("case")): reg["case"] = clean(r.get("case"))
            regel_verbs.append(reg)

    if missing_p2:
        warn.append(f"verbs: unregel без partizip2 ({len(missing_p2)}): {missing_p2[:8]}")
    if missing_aux:
        warn.append(f"verbs: unregel без aux ({len(missing_aux)}): {missing_aux[:8]}")
    return vocab_items, regel_verbs, conjugations

VERB_VOCAB_SCHEMA = [
    ("alt-de", "altDe", "str"),
    ("ru", "ru", "str"), ("alt-ru", "altRu", "str"), ("type", "type", "str"),
    ("modal", "modal", "bool"), ("separable", "separable", "bool"),
    ("prefix", "prefix", "str"), ("reflexive", "reflexive", "bool"),
    ("impersonal", "impersonal", "bool"), ("case", "case", "str"),
    ("aux", "aux", "str"), ("partizip2", "partizip2", "str"),
    ("praeteritum", "praeteritum", "str"),
    ("level", "level", "str"), ("priority", "priority", "int"),
    ("source", "source", "str"),
    ("example_de", "exampleDe", "str"), ("example_ru", "exampleRu", "str"),
    ("quiz_use", "quizUse", "boolT"), ("strict_order", "strictOrder", "boolT"),
    ("warning", "warning", "str"), ("label", "label", "str"), ("note", "note", "str"),
]
def apply_schema_verb(r):
    out = {}
    for col, key, kind in VERB_VOCAB_SCHEMA:
        v = r.get(col)
        if kind == "bool":
            if is_true(v): out[key] = True
        elif kind == "boolT":
            if is_false(v): out[key] = False
        elif kind == "int":
            iv = to_int(v)
            if iv is not None: out[key] = iv
        else:
            cv = clean(v)
            if cv is not None: out[key] = cv
    return out

def process_pronouns(rows):
    items = []
    for r in rows:
        de = clean(r.get("de"))
        if not de: continue
        item = {"de": de, "pos": "pron"}
        item.update(apply_schema(r, "pronouns"))
        attach_common(item, r, "pron")
        items.append(item)
    return items

def process_numbers(rows):
    items = []
    for r in rows:
        de = clean(r.get("de"))
        if not de: continue
        item = {"de": de, "pos": "num"}
        item.update(apply_schema(r, "numbers"))
        rid = clean(r.get("id"))
        if rid: item["id"] = rid
        kind = item.get("kind")
        domen, group = dg(r)
        topics = []
        tk = topic_key(domen, group)
        if tk:
            topics.append(tk)
        else:
            topics.append("nums:" + kind if kind else "nums:basic")
        # буфер «Новые»
        status = new_status(r.get("new"))
        if status:
            item["new"] = True
            if status == "update": item["_newUpdate"] = True
            nt = NEW_TOPIC.get("num")
            if nt and nt not in topics:
                topics.append(nt)
        item["topics"] = topics
        if domen: item["domen"] = domen
        if group: item["group"] = group
        items.append(item)
    return items

def process_sounds(rows):
    items = []
    for r in rows:
        combo = clean(r.get("combo"))
        if not combo: continue
        item = apply_schema(r, "sounds")
        item = {"combo": combo, **{k: v for k, v in item.items() if k != "combo"}}
        rid = clean(r.get("id"))
        if rid: item["id"] = rid
        domen, group = dg(r)
        if domen: item["domen"] = domen
        if group: item["group"] = group
        # буфер «Новые»
        topics = []
        tk = topic_key(domen, group)
        if tk: topics.append(tk)
        status = new_status(r.get("new"))
        if status:
            item["new"] = True
            if status == "update": item["_newUpdate"] = True
            nt = NEW_TOPIC.get("sound")
            if nt and nt not in topics:
                topics.append(nt)
        if topics: item["topics"] = topics
        items.append(item)
    return items

def process_terms(rows):
    items = []
    for r in rows:
        term = clean(r.get("de"))
        if not term: continue
        item = apply_schema(r, "terms")  # de→term внутри схемы
        item = {"term": term, **{k: v for k, v in item.items() if k != "term"}}
        rid = clean(r.get("id"))
        if rid: item["id"] = rid
        domen, group = dg(r)
        tk = topic_key(domen, group)
        if tk: item["topic"] = tk
        if domen: item["domen"] = domen
        if group: item["group"] = group
        # буфер «Новые»
        topics = []
        if tk: topics.append(tk)
        status = new_status(r.get("new"))
        if status:
            item["new"] = True
            if status == "update": item["_newUpdate"] = True
            nt = NEW_TOPIC.get("term")
            if nt and nt not in topics:
                topics.append(nt)
        if topics: item["topics"] = topics
        items.append(item)
    return items

def process_rules(rows, warn):
    items = []
    for r in rows:
        title = clean(r.get("title"))
        if not title: continue
        item = {"title": title}
        item.update(apply_schema(r, "rules"))
        rid = clean(r.get("id"))
        if rid: item["id"] = rid
        domen, group = dg(r)
        tk = topic_key(domen, group)
        if tk:
            item["topic"] = tk
        else:
            warn.append(f"rules: нет topic (пустые domen/group) — «{title[:40]}»")
        if domen: item["domen"] = domen
        if group: item["group"] = group
        # буфер «Новые» — кладём и флаг, и topic (для рендера на вкладке Neu)
        topics = []
        if tk: topics.append(tk)
        status = new_status(r.get("new"))
        if status:
            item["new"] = True
            if status == "update": item["_newUpdate"] = True
            nt = NEW_TOPIC.get("rule")
            if nt and nt not in topics:
                topics.append(nt)
        if topics: item["topics"] = topics
        items.append(item)
    return items

DIFFICULTY_DEFAULTS = {"mc": 1, "fill": 4, "tiles": 4, "conj": 3, "open": 5}
TYPE_ALIASES = {"build": "tiles"}  # движок тренажёра знает только tiles

def process_questions(rows, warn):
    items = []
    no_topic = 0
    for r in rows:
        qtype_raw = clean(r.get("type"))
        q = clean(r.get("q"))
        if not (qtype_raw and q): continue
        qtype = TYPE_ALIASES.get(qtype_raw.lower(), qtype_raw)

        domen, group = dg(r)
        tk = topic_key(domen, group)
        item = {"type": qtype, "q": q}
        if tk:
            item = {"topic": tk, **item}
        else:
            no_topic += 1

        level = clean(r.get("level"))
        if level: item["level"] = level
        diff = to_int(r.get("difficulty"))
        item["difficulty"] = diff if diff is not None else DIFFICULTY_DEFAULTS.get(qtype, 1)

        if qtype == "mc":
            opts = clean(r.get("opts"))
            if opts: item["opts"] = [s.strip() for s in opts.split("|")]
            ans = to_int(r.get("ans_mc"))
            if ans is not None: item["ans"] = ans
        elif qtype == "fill":
            ans = clean(r.get("answer"))
            if ans: item["ans"] = ans
            alt = clean(r.get("altAns"))
            if alt: item["altAns"] = [s.strip() for s in alt.split("|")]
        elif qtype == "tiles":
            words = clean(r.get("words"))
            if words: item["words"] = [s.strip() for s in words.split("|")]
        elif qtype == "conj":
            pr = clean(r.get("pronouns"))
            if pr: item["pronouns"] = [s.strip() for s in pr.split("|")]
            fm = clean(r.get("forms"))
            if fm: item["ans"] = [s.strip() for s in fm.split("|")]
        if clean(r.get("hint")): item["hint"] = clean(r.get("hint"))
        if clean(r.get("explain")): item["explain"] = clean(r.get("explain"))
        rid = clean(r.get("id"))
        if rid: item["id"] = rid
        items.append(item)
    if no_topic:
        warn.append(f"questions: без topic (пустые domen/group): {no_topic} из {len(items)}")
    return items

# ═══════════════════════════════════════════════════════════════
# Settings-XLSX → TAXONOMY (генерически, все колонки)
# ═══════════════════════════════════════════════════════════════
def process_settings(wb):
    """Читает Settings-XLSX в структуру { page: [ {domen, group, ...extra} ] }.
    Читает ВСЕ колонки листа — будущие колонки подхватятся автоматически."""
    name = "Settings-XLSX"
    if name not in wb.sheetnames:
        print(f"  ⚠ Лист '{name}' не найден — TAXONOMY будет пустой")
        return {}
    ws = wb[name]
    headers = [clean(c.value) for c in ws[1]]
    tax = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = {h: clean(v) for h, v in zip(headers, row) if h}
        page = rec.get("page")
        if not page: continue
        entry = {k: v for k, v in rec.items() if k != "page" and v is not None}
        tax.setdefault(page, []).append(entry)
    return tax

def validate_against_taxonomy(items, page, tax, warn):
    """Варнит, если (domen, group) элемента нет в Settings для данной page."""
    if page not in tax: return
    valid = {(e.get("domen"), e.get("group")) for e in tax[page]}
    seen_bad = set()
    for it in items:
        d, g = it.get("domen"), it.get("group")
        if d is None and g is None: continue
        if (d, g) not in valid and (d, g) not in seen_bad:
            seen_bad.add((d, g))
            warn.append(f"{page}: пара ({d}/{g}) отсутствует в Settings-XLSX")

# ═══════════════════════════════════════════════════════════════
# ID: уважаем xlsx, автоген для пустых, детект дублей
# ═══════════════════════════════════════════════════════════════
def assign_ids(items, prefix, warn, label, width=4, warn_blanks=True):
    used = set()
    dupes = set()
    for it in items:
        rid = it.get("id")
        rid = str(rid).strip() if rid not in (None, "") else None
        if rid:
            if rid in used: dupes.add(rid)
            used.add(rid)
    # автоген свободных номеров для пустых
    counter = 1
    def next_free():
        nonlocal counter
        while True:
            cand = f"{prefix}{counter:0{width}d}"
            counter += 1
            if cand not in used:
                used.add(cand)
                return cand
    blanks = 0
    for it in items:
        rid = it.get("id")
        rid = str(rid).strip() if rid not in (None, "") else None
        if not rid:
            rid = next_free()
            blanks += 1
        # id первым ключом
        rest = {k: v for k, v in it.items() if k != "id"}
        it.clear()
        it["id"] = rid
        it.update(rest)
    if dupes:
        warn.append(f"{label}: дубликаты id ({len(dupes)}): {sorted(dupes)[:8]}")
    if blanks and warn_blanks:
        warn.append(f"{label}: пустых id заполнено автогеном: {blanks}")

# ═══════════════════════════════════════════════════════════════
# Сохранение констант из старого data.js
# ═══════════════════════════════════════════════════════════════
def extract_block(content, marker):
    start_marker = f"const {marker} = "
    start = content.find(start_marker)
    if start == -1: return None
    start += len(start_marker)
    if content[start] not in "[{": return None
    open_ch = content[start]
    close_ch = "]" if open_ch == "[" else "}"
    depth = 0
    pos = start
    while pos < len(content):
        c = content[pos]
        if c == open_ch:
            depth += 1
        elif c == close_ch:
            depth -= 1
            if depth == 0:
                return content[start:pos+1]
        elif c == '"':
            pos += 1
            while pos < len(content) and content[pos] != '"':
                pos += 2 if content[pos] == '\\' else 1
        elif c == "'":
            pos += 1
            while pos < len(content) and content[pos] != "'":
                pos += 2 if content[pos] == '\\' else 1
        pos += 1
    return None

# ═══════════════════════════════════════════════════════════════
# ОСНОВНОЙ ПРОЦЕСС
# ═══════════════════════════════════════════════════════════════
if __name__ == '__main__':
    arg_parser = argparse.ArgumentParser(description="Собирает data.js из database.xlsx")
    arg_parser.add_argument("--dry-run", action="store_true",
                             help="Только посчитать, что будет переведено (VOCAB/CONJUGATIONS/"
                                  "REGEL_VERBS/exampleEn) и сколько это символов — без обращения "
                                  "к DeepL и без записи data.js/translations.json")
    arg_parser.add_argument("--force", action="store_true",
                             help="Продолжить перевод VOCAB, даже если он затрагивает больше 20%% словаря")
    arg_parser.add_argument("--force-changelog", action="store_true",
                             help="Перевести записи changelog.js заново, даже если titleEn уже есть. "
                                  "Записи с \"manualEn\": true не трогаются никогда. Без флага "
                                  "переводятся только записи без titleEn.")
    arg_parser.add_argument("--arbiter", choices=["none", "haiku"], default="none",
                             help="Разбор флагов обратного перевода VOCAB.en. 'none' — все флаги в "
                                  "reports/en_review_vocab.md для разбора в чате. 'haiku' — сначала "
                                  f"прогоняет флаги через Claude Haiku ({ARBITER_MODEL}, ANTHROPIC_API_KEY), "
                                  "нерешённое (verdict=skip) всё равно уходит в тот же отчёт.")
    args = arg_parser.parse_args()

    WARN = []

    print(f"Читаю {XLSX_PATH}...")
    wb = load_workbook(XLSX_PATH, data_only=True)

    print("\n=== Settings-XLSX → TAXONOMY ===")
    TAXONOMY = process_settings(wb)
    print(f"  ✓ {sum(len(v) for v in TAXONOMY.values())} пар в {len(TAXONOMY)} страницах")

    print("\n=== nouns ===")
    nouns = process_simple_vocab(read_sheet(wb, "nouns", WARN), "nouns", "noun", WARN)
    print(f"  ✓ {len(nouns)} существительных")

    print("\n=== verbs ===")
    verbs_vocab, regel_verbs, conjugations = process_verbs(read_sheet(wb, "verbs", WARN), WARN)
    print(f"  ✓ {len(verbs_vocab)} в VOCAB, {len(regel_verbs)} regel, {len(conjugations)} unregel")

    print("\n=== ruForms (личные формы наст. времени, pymorphy3) ===")
    if _MORPH is None:
        print("  ⚠ pymorphy3 не установлен — пропускаю (pip install pymorphy3 pymorphy3-dicts-ru)")
        WARN.append("pymorphy3 не установлен — ruForms не сгенерированы, tiles Ур.4 используют сырой v.ru")
    else:
        ru_forms_fail = []
        for v in verbs_vocab:
            forms = generate_ru_forms(v.get("ru"))
            if forms:
                v["ruForms"] = forms
            else:
                ru_forms_fail.append(v.get("ru") or v.get("de"))
        print(f"  ✓ {len(verbs_vocab) - len(ru_forms_fail)}/{len(verbs_vocab)} глаголов проспрягано, "
              f"{len(ru_forms_fail)} — fallback на v.ru")
        if ru_forms_fail:
            WARN.append(f"ruForms: не удалось проспрягать {len(ru_forms_fail)} глаголов "
                        f"(fallback на v.ru, проверить вручную при желании): {ru_forms_fail}")

    print("\n=== adjectives ===")
    adjectives = process_simple_vocab(read_sheet(wb, "adjectives", WARN), "adjectives", "adj", WARN)
    print(f"  ✓ {len(adjectives)} прилагательных")

    print("\n=== adverbs ===")
    adverbs = process_simple_vocab(read_sheet(wb, "adverbs", WARN), "adverbs", "adv", WARN)
    print(f"  ✓ {len(adverbs)} наречий")

    print("\n=== phrases ===")
    phrases = process_simple_vocab(read_sheet(wb, "phrases", WARN), "phrases", "phrase", WARN)
    print(f"  ✓ {len(phrases)} фраз")

    print("\n=== pronouns ===")
    pronouns = process_pronouns(read_sheet(wb, "pronouns", WARN))
    print(f"  ✓ {len(pronouns)} местоимений")

    print("\n=== numbers ===")
    numbers = process_numbers(read_sheet(wb, "numbers", WARN))
    print(f"  ✓ {len(numbers)} чисел")

    print("\n=== terms ===")
    terms = process_terms(read_sheet(wb, "terms", WARN))
    print(f"  ✓ {len(terms)} терминов")

    print("\n=== sounds ===")
    sounds = process_sounds(read_sheet(wb, "sounds", WARN))
    print(f"  ✓ {len(sounds)} звуков")

    print("\n=== rules ===")
    rules = process_rules(read_sheet(wb, "rules", WARN), WARN)
    print(f"  ✓ {len(rules)} правил")

    print("\n=== questions ===")
    questions = process_questions(read_sheet(wb, "questions", WARN), WARN)
    print(f"  ✓ {len(questions)} вопросов")

    # Валидация пар против Settings
    for items, page in [(nouns, "nouns"), (verbs_vocab, "verbs"), (adjectives, "adjectives"),
                        (adverbs, "adverbs"), (phrases, "phrases")]:
        validate_against_taxonomy(items, page, TAXONOMY, WARN)

    # ═══════════════════════════════════════════════════════════════
    # Считываем константы фронта из старого data.js
    # ═══════════════════════════════════════════════════════════════
    print("\n=== Читаю старый data.js (PHRASE_UNITS / SENTENCE_TEMPLATES) ===")
    old_content = OLD_DATA_JS.read_text(encoding="utf-8")
    phrase_units_block = extract_block(old_content, "PHRASE_UNITS")
    sentence_templates_block = extract_block(old_content, "SENTENCE_TEMPLATES")
    for nm, b in [("PHRASE_UNITS", phrase_units_block),
                  ("SENTENCE_TEMPLATES", sentence_templates_block)]:
        if b is None:
            WARN.append(f"не нашёл {nm} в старом data.js — заглушка")

    # ═══════════════════════════════════════════════════════════════
    # Назначаем id
    # ═══════════════════════════════════════════════════════════════
    all_vocab = []
    for items, prefix, lbl in [
        (nouns, "n", "nouns"), (verbs_vocab, "v", "verbs"), (adjectives, "a", "adjectives"),
        (adverbs, "d", "adverbs"), (pronouns, "p", "pronouns"),
        (numbers, "u", "numbers"), (phrases, "f", "phrases"),
    ]:
        assign_ids(items, prefix, WARN, lbl)
        all_vocab.extend(items)

    assign_ids(regel_verbs, "r", WARN, "regel_verbs", width=3, warn_blanks=False)
    assign_ids(conjugations, "c", WARN, "conjugations", width=3, warn_blanks=False)
    assign_ids(questions, "q", WARN, "questions")
    assign_ids(rules, "rl", WARN, "rules", width=3)
    assign_ids(terms, "t", WARN, "terms", width=3)
    assign_ids(sounds, "s", WARN, "sounds", width=3)

    print("\n=== EN-перевод словаря (мультиязычность, слой 2 — DeepL, этап B) ===")
    # translations.json — единственный кэш для всех переводимых полей VOCAB
    # (en/exampleEn/altEn/note/warning) и CONJUGATIONS/REGEL_VERBS, хэш-ключи по
    # источнику. Старый _parse_old_array_by_id (диф по id против предыдущего
    # data.js) удалён на этапе G — это было последнее место, где промах кэша
    # приводил к молчаливому полному перепереводу (id перевыдаётся каждый
    # прогон). Фолбэка на старый data.js больше нет нигде.

    translations = load_translations(WARN)
    load_manual_vocab_skip(all_vocab, translations, WARN)
    vocab_candidates, vocab_cache_hits, vocab_chars = compute_vocab_plan(all_vocab, translations, WARN)
    example_candidates, example_cache_hits, example_chars = \
        compute_example_plan(all_vocab, translations, WARN)
    alt_en_candidates, alt_en_cache_hits, alt_en_flagged_cached, alt_en_skipped, \
        alt_en_chars_de, alt_en_chars_ru = compute_alt_en_plan(all_vocab, translations, WARN)
    conj_candidates, conj_cache_hits, conj_chars = \
        compute_verb_field_plan(conjugations, "verb", translations["CONJUGATIONS"] if translations else {}, WARN)
    regel_candidates, regel_cache_hits, regel_chars = \
        compute_verb_field_plan(regel_verbs, "verb", translations["REGEL_VERBS"] if translations else {}, WARN)

    threshold_ok = check_vocab_threshold(len(vocab_candidates), all_vocab, args.force)

    changelog_entries, changelog_chars = plan_changelog(
        SCRIPT_DIR / "changelog.js", args.force_changelog, WARN)

    if args.dry_run:
        print("\n=== --dry-run: без обращения к API ===")
        print(f"  VOCAB.en (DE→EN + контекст + обратный перевод):")
        print(f"    кэш-хиты (translations.json): {vocab_cache_hits}")
        print(f"    к переводу: {len(vocab_candidates)} слов, {vocab_chars} символов (прямой перевод)")
        print(f"    + обратный перевод (EN→DE) той же длины по каждому кандидату — "
              f"по историческому замеру (докладу в BRIEF §4B) это ещё ~{int(vocab_chars * 1.2)} символов")
        print(f"    порог 20%: {'ПРЕВЫШЕН, нужен --force' if not threshold_ok else 'не превышен'}")
        print(f"  VOCAB.exampleEn (exampleDe, DE→EN): кэш-хитов {example_cache_hits}, "
              f"к переводу {len(example_candidates)} записей, {example_chars} символов")
        print(f"  VOCAB.altEn (этап G — altDe DE→EN / altRu RU→EN + контекст, без обратного перевода):")
        print(f"    кэш-хитов {alt_en_cache_hits} (+{alt_en_flagged_cached} закэш. флагов), "
              f"к переводу {len(alt_en_candidates)} записей "
              f"({alt_en_chars_de} символов altDe + {alt_en_chars_ru} altRu), "
              f"пропущено (нет альт. формы) {alt_en_skipped}")
        print(f"  CONJUGATIONS.en (verb, DE→EN): кэш-хитов {conj_cache_hits}, "
              f"к переводу {len(conj_candidates)} записей, {conj_chars} символов")
        print(f"  REGEL_VERBS.en (verb, DE→EN): кэш-хитов {regel_cache_hits}, "
              f"к переводу {len(regel_candidates)} записей, {regel_chars} символов")
        print(f"  RULES: этап C, не входит в --dry-run этапа B (символы не посчитаны)")
        print(f"  TERMS/SOUNDS: только manual, не трогаются")
        print(f"  changelog.js (этап F — ignore-теги, RU→EN, построчно):")
        print(f"    к переводу: {len(changelog_entries)} записей"
              f"{' [--force-changelog]' if args.force_changelog else ''}")
        print(f"    ~{changelog_chars} символов уйдёт в DeepL "
              f"(только строки с кириллицей; латиница/числа копируются без API)")
        if WARN:
            print(f"\n⚠ Предупреждения ({len(WARN)}):")
            for w in WARN:
                print("   •", w)
        print("\n(--dry-run: data.js и translations.json не изменены, к DeepL не обращались)")
        sys.exit(0)

    if not threshold_ok:
        sys.exit(1)

    vocab_report_rows = []
    alt_en_report_rows = []
    alt_en_n_de = alt_en_n_ru = 0
    usage_before, usage_limit = _deepl_usage(DEEPL_API_KEY)
    if usage_before is not None:
        print(f"  DeepL /usage до прогона: {usage_before:,}/{usage_limit:,} символов")
    if translations is not None:
        vocab_report_rows, _vocab_ok = run_vocab_translation(vocab_candidates, translations["VOCAB"], WARN)
        run_example_translation(example_candidates, translations["VOCAB"], WARN)
        alt_en_report_rows, alt_en_n_de, alt_en_n_ru = \
            run_alt_en_translation(alt_en_candidates, translations["VOCAB"], WARN)
        run_verb_field_translation(conj_candidates, "verb", translations["CONJUGATIONS"], WARN, "CONJUGATIONS")
        run_verb_field_translation(regel_candidates, "verb", translations["REGEL_VERBS"], WARN, "REGEL_VERBS")
        if args.arbiter == "haiku":
            vocab_report_rows, arbiter_stats = run_haiku_arbiter(vocab_report_rows, translations["VOCAB"], WARN)
            arbiter_report_path = write_arbiter_report(arbiter_stats)
            print(f"  ✓ отчёт арбитра: {arbiter_report_path}")
        save_translations(translations)
    report_path = write_vocab_review_report(vocab_report_rows)
    print(f"  ✓ отчёт: {report_path}")
    alt_en_report_path = write_alt_en_report(
        alt_en_report_rows, alt_en_n_de, alt_en_n_ru, alt_en_skipped,
        alt_en_cache_hits, alt_en_flagged_cached,
        translations["VOCAB"] if translations is not None else {})
    print(f"  ✓ отчёт: {alt_en_report_path}")

    # RULES — этап C: гибрид (31 ручное исключение + 28 через ignore-теги).
    # Кэш — translations.json["RULES"], не старый data.js (см. докстринг
    # блока выше).
    print("\n=== EN-перевод RULES (этап C — исключения + ignore-теги) ===")
    notes_report_rows = []
    if translations is not None:
        load_manual_rules_batches(translations, WARN)
        rules_manual_ids = apply_manual_rules(rules, translations, WARN)
        print(f"  ✓ RULES manual применено: {len(rules_manual_ids)}/{len(rules)}")
        rules_report_rows = []
        for field, en_field in (("title", "titleEn"), ("content_md", "content_md_en"), ("note", "noteEn")):
            candidates = compute_ignoretag_candidates(rules, field, en_field, translations["RULES"],
                                                       lambda r: r["id"], rules_manual_ids)
            rules_report_rows += translate_ignoretag_field(candidates, field, en_field, translations["RULES"],
                                                             lambda r: r["id"], WARN, "RULES")
        rules_report_path = write_rules_review_report(rules, translations, rules_report_rows)
        print(f"  ✓ отчёт: {rules_report_path}")
    else:
        WARN.append("translations.json недоступен — RULES.titleEn/content_md_en/noteEn не обновлены в этом прогоне")

    # VOCAB.note/warning — этап D: те же ignore-теги, что для RULES (349/437
    # примечаний содержат немецкий — сырой RU→EN их портил так же, как правила).
    # Кэш — тот же VOCAB[hash], что у основного слова/exampleEn (этап B).
    print("\n=== EN-перевод VOCAB.note/warning (этап D — ignore-теги) ===")
    if translations is not None:
        vcache = translations["VOCAB"]
        for field, en_field in (("note", "noteEn"), ("warning", "warningEn")):
            candidates = compute_ignoretag_candidates(all_vocab, field, en_field, vcache, vocab_hash)
            notes_report_rows += translate_ignoretag_field(candidates, field, en_field, vcache,
                                                             vocab_hash, WARN, "VOCAB")
    else:
        WARN.append("translations.json недоступен — VOCAB.noteEn/warningEn не обновлены в этом прогоне")

    # TERMS/SOUNDS — этап E: автоперевод убран, значения только из
    # translations-terms-sounds.json (source: manual). TERMS.note — этап D
    # (ignore-теги), но фактически все 54 термина уже приходят с noteEn
    # готовым в том же manual-файле, так что кандидатов для DeepL не будет,
    # если файл полон — механизм всё равно оставлен как страховка на будущее.
    print("\n=== EN-перевод TERMS/SOUNDS (этап E — только manual) ===")
    if translations is not None:
        load_manual_terms_sounds(translations, WARN)
        apply_manual_terms(terms, translations, WARN)
        apply_manual_sounds(sounds, translations, WARN)
        tcandidates = compute_ignoretag_candidates(terms, "note", "noteEn", translations["TERMS"],
                                                    lambda t: t["id"])
        notes_report_rows += translate_ignoretag_field(tcandidates, "note", "noteEn", translations["TERMS"],
                                                         lambda t: t["id"], WARN, "TERMS")
        save_translations(translations)
    else:
        WARN.append("translations.json недоступен — TERMS/SOUNDS.en не обновлены в этом прогоне")

    notes_report_path = write_notes_review_report(notes_report_rows)
    print(f"  ✓ отчёт: {notes_report_path}")

    print("\n=== EN-перевод changelog.js ===")
    translate_changelog(SCRIPT_DIR / "changelog.js", WARN, args.force_changelog)

    print("\n=== Генерирую BLOCKS / TOPIC_TITLES / TAB_TITLES (немецкий) ===")
    BLOCKS_DATA = build_blocks(TAXONOMY, all_vocab)
    TOPIC_TITLES_DATA = build_topic_titles(TAXONOMY)
    TAB_TITLES_DATA = build_tab_titles()
    TOPIC_TITLES_RU_DATA, TAB_TITLES_RU_DATA = build_ru_titles(TAXONOMY)
    print(f"  ✓ BLOCKS: {len(BLOCKS_DATA)} блоков, "
          f"{sum(len(b.get('subblocks',[])) for b in BLOCKS_DATA)} подблоков")
    print(f"  ✓ TOPIC_TITLES: {len(TOPIC_TITLES_DATA)} ключей, RU: {len(TOPIC_TITLES_RU_DATA)}")

    # ═══════════════════════════════════════════════════════════════
    # Сериализация
    # ═══════════════════════════════════════════════════════════════
    print("\n=== Собираю data.js ===")

    VOCAB_KEYS = ["id", "de", "altDe", "ru", "altRu", "altEn", "en", "pos", "gender", "plural", "altPlural",
                  "level", "topics", "domen", "group", "note", "noteEn", "new",
                  "comparative", "superlative", "antonym", "derivedFrom",
                  "kind", "case", "digit", "transcription", "context",
                  "type", "modal", "separable", "prefix", "reflexive", "impersonal",
                  "aux", "partizip2", "praeteritum", "ruForms",
                  "priority", "source", "exampleDe", "exampleRu", "exampleEn", "quizUse",
                  "strictOrder", "warning", "warningEn", "label"]
    REGEL_KEYS = ["id", "verb", "ru", "en", "forms", "partizip2", "aux", "praeteritum",
                  "separable", "reflexive", "impersonal", "case"]
    CONJ_KEYS = ["id", "verb", "ru", "en", "tense", "modal", "level", "pronouns", "forms",
                 "partizip2", "aux", "praeteritum", "separable", "reflexive",
                 "impersonal", "case"]
    Q_KEYS = ["id", "topic", "level", "difficulty", "type", "q",
              "opts", "ans", "words", "pronouns", "altAns", "hint", "explain"]
    R_KEYS = ["id", "title", "titleEn", "topic", "domen", "group", "level", "content_md", "content_md_en",
              "examples", "examplesEn", "note", "noteEn", "new"]
    T_KEYS = ["id", "term", "plural", "ru", "en", "topic", "domen", "group", "level",
              "priority", "source", "note", "noteEn"]
    S_KEYS = ["id", "combo", "comboEn", "pronunciation", "pronunciationEn", "example", "exampleEn",
              "translation", "domen", "group", "note"]

    out = []
    out.append("// ═══════════════════════════════════════════════════════════════")
    out.append("// data.js — единый источник правды для справочника и тренажёра")
    out.append("// Собрано автоматически из database.xlsx (см. update_data.py)")
    out.append("// ═══════════════════════════════════════════════════════════════")
    out.append(f"// VOCAB: {len(all_vocab)} | REGEL: {len(regel_verbs)} | CONJ: {len(conjugations)}")
    out.append(f"// QUESTIONS: {len(questions)} | RULES: {len(rules)} | TERMS: {len(terms)} | SOUNDS: {len(sounds)}")
    out.append("")

    vocab_by_pos = {}
    for v in all_vocab:
        vocab_by_pos.setdefault(v["pos"], []).append(v)
    out.append("const VOCAB = [")
    POS_ORDER = ["num", "noun", "verb", "adj", "adv", "pron", "phrase"]
    POS_LABEL = {"num": "Числа", "noun": "Существительные", "verb": "Глаголы",
                 "adj": "Прилагательные", "adv": "Наречия", "pron": "Местоимения",
                 "phrase": "Фразы"}
    for pos in POS_ORDER:
        if pos in vocab_by_pos:
            out.append(f"  // ─── {POS_LABEL[pos]} ({len(vocab_by_pos[pos])}) ─────────────")
            for item in vocab_by_pos[pos]:
                out.append(item_to_js(item, VOCAB_KEYS))
    out.append("];")
    out.append("")

    for const_name, data, keys in [
        ("CONJUGATIONS", conjugations, CONJ_KEYS),
        ("REGEL_VERBS", regel_verbs, REGEL_KEYS),
        ("QUESTIONS", questions, Q_KEYS),
        ("SOUNDS", sounds, S_KEYS),
        ("TERMS", terms, T_KEYS),
        ("RULES", rules, R_KEYS),
    ]:
        out.append(f"const {const_name} = [")
        for it in data:
            out.append(item_to_js(it, keys))
        out.append("];")
        out.append("")

    # TAXONOMY (из Settings-XLSX)
    out.append("const TAXONOMY = " + js_value(TAXONOMY) + ";")
    out.append("")

    # ICON_SVGS — единственный источник SVG-иконок (потребители: cheatsheet.html, trainer.html)
    out.append("const ICON_SVGS = " + js_value(ICON_LIBRARY) + ";")
    out.append("")

    # BLOCKS / TOPIC_TITLES / TAB_TITLES — сгенерированы (немецкий, новая таксономия)
    out.append("const BLOCKS = " + js_value(BLOCKS_DATA) + ";")
    out.append("")
    out.append("const TOPIC_TITLES = " + js_value(TOPIC_TITLES_DATA) + ";")
    out.append("")
    out.append("const TAB_TITLES = " + js_value(TAB_TITLES_DATA) + ";")
    out.append("")
    out.append("const TOPIC_TITLES_RU = " + js_value(TOPIC_TITLES_RU_DATA) + ";")
    out.append("")
    out.append("const TAB_TITLES_RU = " + js_value(TAB_TITLES_RU_DATA) + ";")
    out.append("")

    # PHRASE_UNITS — сохранили из старого data.js (если был)
    if phrase_units_block:
        out.append("const PHRASE_UNITS = " + phrase_units_block + ";")
        out.append("")
    if sentence_templates_block:
        out.append("const SENTENCE_TEMPLATES = " + sentence_templates_block + ";")
    else:
        out.append("const SENTENCE_TEMPLATES = [];")
    out.append("")



    OUT_DATA_JS.write_text("\n".join(out), encoding="utf-8")

    # ═══════════════════════════════════════════════════════════════
    # Отчёт
    # ═══════════════════════════════════════════════════════════════
    print(f"\n✓ Записано: {OUT_DATA_JS}")
    print(f"  размер: {OUT_DATA_JS.stat().st_size:,} байт")

    # ═══════════════════════════════════════════════════════════════
    # Что нового (по флагу new=TRUE в xlsx, текущее состояние — не диф
    # с прошлым прогоном). Снять флаг в xlsx после публикации — сам
    # скрипт его не сбрасывает.
    # ═══════════════════════════════════════════════════════════════
    NEW_POS_LABELS = {
        "noun":   ("Nomen", "Nomen"),
        "verb":   ("Verb", "Verben"),
        "adj":    ("Adjektiv", "Adjektive"),
        "adv":    ("Adverb", "Adverbien"),
        "pron":   ("Pronomen", "Pronomen"),
        "num":    ("Zahl", "Zahlen"),
        "phrase": ("Redewendung", "Redewendungen"),
        "term":   ("Begriff", "Begriffe"),
    }
    NEW_POS_ORDER = ["noun", "verb", "adj", "adv", "pron", "num", "phrase", "term"]

    new_counts = {}
    updated_words_total = 0
    for v in all_vocab:
        if v.get("new"):
            new_counts[v["pos"]] = new_counts.get(v["pos"], 0) + 1
            if v.get("_newUpdate"):
                updated_words_total += 1
    new_terms_count = sum(1 for t in terms if t.get("new"))
    if new_terms_count:
        new_counts["term"] = new_terms_count
        updated_words_total += sum(1 for t in terms if t.get("_newUpdate"))

    total_new_words = sum(new_counts.values())
    if total_new_words:
        suffix = f" (davon {updated_words_total} aktualisiert)" if updated_words_total else ""
        print(f"\n+{total_new_words} neue Wörter{suffix}:")
        for pos in NEW_POS_ORDER:
            n = new_counts.get(pos, 0)
            if not n:
                continue
            sg, pl = NEW_POS_LABELS[pos]
            print(f"  · {n} {sg if n == 1 else pl}")

    new_rules = [r for r in rules if r.get("new")]
    if new_rules:
        print(f"\n+{len(new_rules)} neue Regeln:")
        for r in new_rules:
            mark = " ↻ aktualisiert" if r.get("_newUpdate") else ""
            print(f"   • {r['title']}{mark}")

    usage_after, _lim = _deepl_usage(DEEPL_API_KEY)
    if usage_after is not None:
        delta = f" (Δ {usage_after - usage_before:+,})" if usage_before is not None else ""
        print(f"\nDeepL /usage после прогона: {usage_after:,}/{_lim:,} символов{delta}")

    if WARN:
        print(f"\n⚠ Предупреждения ({len(WARN)}):")
        for w in WARN:
            print("   •", w)
    else:
        print("\n✓ Без предупреждений")
